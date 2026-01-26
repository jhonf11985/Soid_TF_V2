from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.db.models import Q, Count
from django.contrib import messages
from django.urls import reverse   # necesario para eliminar_familiar
from datetime import date, timedelta
from django.contrib.auth.decorators import login_required 
from django.utils import timezone
from django.db.models.functions import ExtractDay
from django.http import HttpResponse
from django.conf import settings
from .forms import EnviarFichaMiembroEmailForm
from .forms import EnviarFichaMiembroEmailForm
import tempfile
import subprocess   # ← ESTE
from .models import Miembro, MiembroRelacion, RazonSalidaMiembro, sync_familia_inteligente_por_relacion

from .forms import MiembroForm, MiembroRelacionForm,NuevoCreyenteForm,MiembroSalidaForm, MiembroReingresoForm
import platform   # ← ESTE FALTABA
from core.utils_config import get_edad_minima_miembro_oficial
from django.http import HttpResponse
import traceback
from io import BytesIO
from django.template.loader import render_to_string
import shutil       
import os
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from core.utils_chrome import get_chrome_executable
# Configuración de Chrome/Chromium (ruta opcional)
from openpyxl import Workbook
from datetime import date, timedelta, datetime
from openpyxl import Workbook, load_workbook
from notificaciones_app.utils import crear_notificacion
from finanzas_app.models import MovimientoFinanciero
from django.db.models import Sum
from .forms import MiembroSalidaForm
from django.apps import apps
from core.models import Module
from nuevo_creyente_app.models import NuevoCreyenteExpediente
from django.db.models import Exists, OuterRef
from django.http import JsonResponse
CHROME_PATH = getattr(settings, "CHROME_PATH", None) or os.environ.get("CHROME_PATH")
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import Miembro
from nuevo_creyente_app.models import NuevoCreyenteExpediente
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import re
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.views.decorators.http import require_POST, require_GET
from django.views.generic import DetailView, UpdateView
from django.views.decorators.http import require_http_methods
from urllib.parse import quote




@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_finanzas_desbloquear(request, pk):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "No autorizado"}, status=403)

    password = request.POST.get("password", "").strip()

    if request.user.check_password(password):
        request.session[f"miembro_finanzas_{pk}"] = True
        return JsonResponse({"ok": True})

    return JsonResponse({"ok": False, "error": "Contraseña incorrecta"})


@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_finanzas_bloquear(request, pk):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "No autorizado"}, status=403)

    request.session.pop(f"miembro_finanzas_{pk}", None)
    return JsonResponse({"ok": True})


@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_privado_desbloquear(request, pk):
    # ✅ Solo admins (ajusta si quieres: is_staff en vez de is_superuser)
    if not (request.user.is_staff or request.user.is_superuser):

        return JsonResponse({"ok": False, "error": "No autorizado"}, status=403)

    password = request.POST.get("password", "").strip()

    # ✅ Validar la contraseña del admin logueado
    if request.user.check_password(password):
        request.session[f"miembro_privado_{pk}"] = True
        return JsonResponse({"ok": True})

    return JsonResponse({"ok": False, "error": "Contraseña incorrecta"})

@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_privado_bloquear(request, pk):
    # admin iglesia o superadmin
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "No autorizado"}, status=403)

    request.session.pop(f"miembro_privado_{pk}", None)
    return JsonResponse({"ok": True})


@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def enviar_a_nuevo_creyente(request, pk):
    miembro = get_object_or_404(Miembro, pk=pk)

    # Mantenerte en la misma lista (con filtros) si mandamos "next"
    next_url = request.POST.get("next") or request.GET.get("next")

    # Si no hay next, lo enviamos al módulo Nuevo Creyente
    if not next_url:
        # Ajusta este name si tu módulo usa otro
        # (si tu ruta es /nuevo-creyente/ normalmente es "nuevo_creyente_app:dashboard" o "nuevo_creyente_app:lista")
        next_url = reverse("nuevo_creyente_app:dashboard")
    # 🚫 No permitir enviar si está inactivo
    if not miembro.activo:
        messages.error(
            request,
            "No se puede enviar a seguimiento: este nuevo creyente ya fue dado de salida."
        )
        return redirect(next_url)

    # Validación: ya existe expediente
    if hasattr(miembro, "expediente_nuevo_creyente"):
        messages.info(request, "Este miembro ya fue enviado al módulo de Nuevo Creyente.")
        return redirect(next_url)

    # Crear expediente
    NuevoCreyenteExpediente.objects.create(
        miembro=miembro,
        responsable=request.user
    )

    # Marcar bandera informativa
    if not miembro.nuevo_creyente:
        miembro.nuevo_creyente = True
        miembro.save(update_fields=["nuevo_creyente"])

    messages.success(request, "Enviado correctamente al módulo de Nuevo Creyente.")
    return redirect(next_url)


def _modulo_estructura_activo():
    return Module.objects.filter(
        is_enabled=True,
        code__in=["Estructura", "Unidad", "Unidades"]
    ).exists()
def _modulo_nuevo_creyente_activo():
    return Module.objects.filter(
        is_enabled=True,
        code="nuevo_creyente"
    ).exists()

def _safe_get_model(app_label, model_name):
    try:
        return apps.get_model(app_label, model_name)
    except Exception:
        return None

# -------------------------------------
# DASHBOARD
# -------------------------------------
@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def miembros_dashboard(request):
    miembros = Miembro.objects.filter(activo=True)


    # Edad mínima oficial desde parámetros
    edad_minima = get_edad_minima_miembro_oficial()

    # -----------------------------
    # Cálculo de edades
    # -----------------------------
    def calcular_edad(fecha_nacimiento):
        if not fecha_nacimiento:
            return None
        hoy = date.today()
        edad = hoy.year - fecha_nacimiento.year
        if (hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day):
            edad -= 1
        return edad

    # -----------------------------
    # Conteo de membresía oficial y base de porcentaje
    # (solo >= edad_minima)
    # -----------------------------
    activos = 0
    pasivos = 0
    descarriados = 0
    observacion = 0
    disciplina = 0
    catecumenos = 0  # no bautizados en edad oficial

    # 1) Contamos estados pastorales SOLO para los miembros que siguen activos en la iglesia
    for m in miembros:
        edad = calcular_edad(m.fecha_nacimiento)
        if edad is None or edad < edad_minima:
            # No se considera miembro oficial (niños fuera de la membresía oficial)
            continue

        # Excluir del cálculo de estado pastoral a los nuevos creyentes
        if m.nuevo_creyente:
            continue

        # Si no está bautizado/confirmado, se cuenta como catecúmeno
        if not m.bautizado_confirmado:
            catecumenos += 1
            continue

        # Miembro oficial bautizado: se distribuye por estado
        if m.estado_miembro == "activo":
            activos += 1
        elif m.estado_miembro == "pasivo":
            pasivos += 1
        elif m.estado_miembro == "observacion":
            observacion += 1
        elif m.estado_miembro == "disciplina":
            disciplina += 1
        # 👇 OJO: ya no usamos estado_miembro == "descarriado" aquí;
        #        los descarriados se toman de los miembros INACTIVOS con razón de salida.

    # 2) Recalcular descarriados: miembros inactivos con razón de salida "descarriado"
    miembros_descarriados = Miembro.objects.filter(
        activo=False,
        razon_salida__isnull=False,
    )

    descarriados = 0
    for m in miembros_descarriados:
        edad = calcular_edad(m.fecha_nacimiento)
        if edad is None or edad < edad_minima:
            continue

        # Convertimos la razón de salida a texto y buscamos "descarri"
        texto_razon = str(m.razon_salida).lower()
        if "descarri" in texto_razon:
            descarriados += 1

    # Total de membresía oficial (para referencia general)
    # 👉 Aquí ya NO incluimos a los descarriados porque son miembros inactivos.
    total_oficiales = (
        activos
        + pasivos
        + observacion
        + disciplina
        + catecumenos
    )

    # Usamos el total oficial como base para los porcentajes del panel
    total_base = total_oficiales

    def porcentaje(cantidad, base=None):
        if base is None:
            base = total_base
        if base == 0:
            return 0
        return round((cantidad * 100) / base, 1)


    # -----------------------------
    # Totales generales y distribución por etapa de vida
    # -----------------------------
    total_miembros = miembros.count()

    campo_categoria = Miembro._meta.get_field("categoria_edad")
    choices_dict = dict(campo_categoria.flatchoices)

    distribucion_raw = (
        miembros.values("categoria_edad")
        .exclude(categoria_edad="")
        .annotate(cantidad=Count("id"))
        .order_by("categoria_edad")
    )

    distribucion_etapa_vida = [
        {
            "codigo": row["categoria_edad"],
            "nombre": choices_dict.get(row["categoria_edad"], "Sin definir"),
            "cantidad": row["cantidad"],
        }
        for row in distribucion_raw
    ]

    # -----------------------------
    # Próximos cumpleaños (30 días)
    # -----------------------------
    hoy = date.today()
    fin_rango = hoy + timedelta(days=30)

    cumple_qs = miembros.filter(fecha_nacimiento__isnull=False)

    proximos_cumpleanos = []
    for m in cumple_qs:
        fn = m.fecha_nacimiento
        proximo = fn.replace(year=hoy.year)
        # Si ya pasó este año, usamos el año siguiente
        if proximo < hoy:
            proximo = proximo.replace(year=hoy.year + 1)

        if hoy <= proximo <= fin_rango:
            edad_que_cumple = proximo.year - fn.year
            proximos_cumpleanos.append(
                {
                    "nombre": f"{m.nombres} {m.apellidos}",
                    "fecha": proximo,
                    "edad": edad_que_cumple,
                }
            )

    proximos_cumpleanos = sorted(proximos_cumpleanos, key=lambda x: x["fecha"])

    # -----------------------------
    # KPI específicos
    # -----------------------------
    # Nuevos miembros del mes (según fecha de ingreso)
    nuevos_mes = miembros.filter(
        fecha_ingreso_iglesia__year=hoy.year,
        fecha_ingreso_iglesia__month=hoy.month,
    ).count()

    # Nuevos creyentes en los últimos 7 días
    hace_7_dias = timezone.now() - timedelta(days=7)
    nuevos_creyentes_semana = Miembro.objects.filter(
        nuevo_creyente=True,
        activo=True,              # 👈 CLAVE
        fecha_creacion__gte=hace_7_dias,
    ).count()

    # Miembros recientes (por si lo usamos luego)
    try:
        miembros_recientes = miembros.order_by(
            "-fecha_ingreso_iglesia", "-fecha_creacion"
        )[:5]
    except Exception:
        # Por si no existiera fecha_creacion en el modelo
        miembros_recientes = miembros.order_by("-fecha_ingreso_iglesia", "-id")[:5]

    # -----------------------------
    # Alertas de datos incompletos
    # -----------------------------
    sin_contacto = miembros.filter(
        (Q(telefono__isnull=True) | Q(telefono="")),
        (Q(telefono_secundario__isnull=True) | Q(telefono_secundario="")),
        (Q(email__isnull=True) | Q(email="")),
    ).count()

    sin_foto = miembros.filter(
        Q(foto__isnull=True) | Q(foto="")
    ).count()

    sin_fecha_nacimiento = miembros.filter(
        fecha_nacimiento__isnull=True
    ).count()

    # -----------------------------
    # Últimas salidas / traslados
    # -----------------------------
    ultimas_salidas = (
        Miembro.objects.filter(activo=False, fecha_salida__isnull=False)
        .order_by("-fecha_salida", "apellidos", "nombres")[:5]
    )

    # -----------------------------
    # Nuevos creyentes recientes (máx. 5)
    # -----------------------------
    nuevos_creyentes_recientes = (
        Miembro.objects.filter(nuevo_creyente=True)
        .order_by("-fecha_creacion", "-id")[:5]
    )

    context = {
        "titulo_pagina": "Miembros",
        "descripcion_pagina": (
            f"Resumen de la membresía oficial "
            f"(mayores de {edad_minima} años) y distribución general."
        ),
        # KPI
        "total_miembros": total_miembros,
        "total_oficiales": total_oficiales,
        "nuevos_mes": nuevos_mes,
        "nuevos_creyentes_semana": nuevos_creyentes_semana,
        # Tarjetas oficiales
        "activos": activos,
        "pasivos": pasivos,
        "descarriados": descarriados,
        "observacion": observacion,
        "disciplina": disciplina,
        "catecumenos": catecumenos,
        "pct_activos": porcentaje(activos),
        "pct_pasivos": porcentaje(pasivos),
        "pct_descarriados": porcentaje(descarriados),
        "pct_observacion": porcentaje(observacion),
        "pct_catecumenos": porcentaje(catecumenos),
        "pct_disciplina": porcentaje(disciplina),

        # Gráfico y listas
        "distribucion_etapa_vida": distribucion_etapa_vida,
        "proximos_cumpleanos": proximos_cumpleanos,
        "miembros_recientes": miembros_recientes,
        # Alertas
        "sin_contacto": sin_contacto,
        "sin_foto": sin_foto,
        "sin_fecha_nacimiento": sin_fecha_nacimiento,
        "ultimas_salidas": ultimas_salidas,
        # Panel nuevos creyentes
        "nuevos_creyentes_recientes": nuevos_creyentes_recientes,
        # Parámetro global para la vista
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }
    return render(request, "miembros_app/miembros_dashboard.html", context)

# -------------------------------------
# FUNCIÓN AUXILIAR DE FILTRO DE MIEMBROS
# -------------------------------------
# =====================================================================
# ACTUALIZACIÓN DE LA FUNCIÓN filtrar_miembros en views.py
# Agregar estos filtros: rol_ministerial, estado_ministerial, tiene_credenciales
# =====================================================================

# REEMPLAZAR la función filtrar_miembros completa con esta versión:

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def filtrar_miembros(request, miembros_base):
    """
    Aplica todos los filtros del listado general de miembros.
    Devuelve:
        - miembros (queryset o lista filtrada)
        - filtros_context (diccionario para la plantilla)
    """

    # -----------------------------
    # 1. Leer parámetros del GET
    # -----------------------------
    query = request.GET.get("q", "").strip()

    estado = request.GET.get("estado", "").strip()
    categoria_edad_filtro = request.GET.get("categoria_edad", "").strip()
    genero = request.GET.get("genero", "").strip()

    # En el HTML: bautizado = "" / "1" / "0"
    bautizado = request.GET.get("bautizado", "").strip()

    # ═══════════════════════════════════════════════════════════════════
    # NUEVOS FILTROS MINISTERIALES
    # ═══════════════════════════════════════════════════════════════════
    rol_ministerial = request.GET.get("rol_ministerial", "").strip()
    estado_ministerial = request.GET.get("estado_ministerial", "").strip()
    # tiene_credenciales: "" / "1" / "0"
    tiene_credenciales_filtro = request.GET.get("tiene_credenciales", "").strip()

    # Checkboxes (vienen solo si están marcados)
    tiene_contacto = request.GET.get("tiene_contacto", "") == "1"
    mostrar_todos = request.GET.get("mostrar_todos", "") == "1"
    incluir_ninos = request.GET.get("incluir_ninos", "") == "1"

    usar_rango_edad = request.GET.get("usar_rango_edad", "") == "1"
    edad_min_str = request.GET.get("edad_min", "").strip()
    edad_max_str = request.GET.get("edad_max", "").strip()

    edad_min = None
    edad_max = None
    if edad_min_str:
        try:
            edad_min = int(edad_min_str)
        except ValueError:
            edad_min = None
    if edad_max_str:
        try:
            edad_max = int(edad_max_str)
        except ValueError:
            edad_max = None

    hoy = date.today()

    # -----------------------------
    # 2. Base de miembros
    # -----------------------------
    miembros = miembros_base

    # Solo activos por defecto. Si NO marcas "Mostrar todos",
    # se filtra por activo=True
    if not mostrar_todos:
        miembros = miembros.filter(activo=True)

    # -----------------------------
    # 3. Exclusión de niños por defecto
    # (niños = menores de 12 años)
    # -----------------------------
    CORTE_NINOS = 12
    cutoff_ninos = hoy - timedelta(days=CORTE_NINOS * 365)

    # Categorías que consideramos "de niños"
    categorias_nino = ("infante", "nino")

    # Si NO marcamos "incluir_ninos" y tampoco estamos filtrando
    # por una categoría de niños, entonces ocultamos los < 12 años
    if not incluir_ninos and categoria_edad_filtro not in categorias_nino:
        miembros = miembros.filter(
            Q(fecha_nacimiento__lte=cutoff_ninos) | Q(fecha_nacimiento__isnull=True)
        )

    # -----------------------------
    # 4. Búsqueda general
    # -----------------------------
    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query)
            | Q(apellidos__icontains=query)
            | Q(telefono__icontains=query)
            | Q(email__icontains=query)
            | Q(cedula__icontains=query)
        )

    # -----------------------------
    # 5. Filtros simples
    # -----------------------------
    if estado:
        miembros = miembros.filter(estado_miembro=estado)

    if genero:
        miembros = miembros.filter(genero=genero)

    if categoria_edad_filtro:
        miembros = miembros.filter(categoria_edad=categoria_edad_filtro)

    # Bautismo: en el HTML usas 1/0
    if bautizado == "1":
        # Solo bautizados
        miembros = miembros.filter(bautizado_confirmado=True)
    elif bautizado == "0":
        # No bautizados (o sin dato)
        miembros = miembros.filter(
            Q(bautizado_confirmado=False) | Q(bautizado_confirmado__isnull=True)
        )

    # ═══════════════════════════════════════════════════════════════════
    # NUEVOS FILTROS MINISTERIALES
    # ═══════════════════════════════════════════════════════════════════
    if rol_ministerial:
        miembros = miembros.filter(rol_ministerial=rol_ministerial)

    if estado_ministerial:
        miembros = miembros.filter(estado_ministerial=estado_ministerial)

    if tiene_credenciales_filtro == "1":
        miembros = miembros.filter(tiene_credenciales=True)
    elif tiene_credenciales_filtro == "0":
        miembros = miembros.filter(tiene_credenciales=False)

    # Solo con contacto
    if tiene_contacto:
        miembros = miembros.filter(
            Q(telefono__isnull=False, telefono__gt="")
            | Q(telefono_secundario__isnull=False, telefono_secundario__gt="")
            | Q(email__isnull=False, email__gt="")
        )

    # Orden base
    miembros = miembros.order_by("nombres", "apellidos")

    # -----------------------------
    # 6. Filtro por rango de edad
    # (solo si el check está marcado)
    # -----------------------------
    if usar_rango_edad and (edad_min is not None or edad_max is not None):
        miembros_filtrados = []

        for m in miembros:
            if not m.fecha_nacimiento:
                # Si no tiene fecha de nacimiento, no se puede filtrar por edad
                continue

            # Usamos el método del modelo si existe
            if hasattr(m, "calcular_edad"):
                edad = m.calcular_edad()
            else:
                fn = m.fecha_nacimiento
                edad = hoy.year - fn.year
                if (hoy.month, hoy.day) < (fn.month, fn.day):
                    edad -= 1

            if edad is None:
                continue

            if edad_min is not None and edad < edad_min:
                continue
            if edad_max is not None and edad > edad_max:
                continue

            miembros_filtrados.append(m)

        # Ahora miembros es una lista, no un queryset
        miembros = miembros_filtrados

    # -----------------------------
    # 7. Choices para los selects
    # -----------------------------
    campo_estado = Miembro._meta.get_field("estado_miembro")
    estados_choices = list(campo_estado.flatchoices)

    campo_categoria = Miembro._meta.get_field("categoria_edad")
    categorias_choices = list(campo_categoria.flatchoices)

    try:
        campo_genero = Miembro._meta.get_field("genero")
        generos_choices = list(campo_genero.flatchoices)
    except Exception:
        generos_choices = []

    # ═══════════════════════════════════════════════════════════════════
    # CHOICES PARA NUEVOS FILTROS MINISTERIALES
    # ═══════════════════════════════════════════════════════════════════
    try:
        campo_rol_ministerial = Miembro._meta.get_field("rol_ministerial")
        roles_ministeriales_choices = [
            (k, v) for k, v in campo_rol_ministerial.flatchoices if k
        ]
    except Exception:
        roles_ministeriales_choices = []

    try:
        campo_estado_ministerial = Miembro._meta.get_field("estado_ministerial")
        estados_ministeriales_choices = [
            (k, v) for k, v in campo_estado_ministerial.flatchoices if k
        ]
    except Exception:
        estados_ministeriales_choices = []

    # -----------------------------
    # 8. Contexto de filtros
    # -----------------------------
    filtros_context = {
        "query": query,
        "mostrar_todos": mostrar_todos,
        "incluir_ninos": incluir_ninos,
        "estado": estado,
        "categoria_edad_filtro": categoria_edad_filtro,
        "genero_filtro": genero,
        "bautizado": bautizado,
        "tiene_contacto": tiene_contacto,
        "estados_choices": estados_choices,
        "categorias_choices": categorias_choices,
        "generos_choices": generos_choices,
        "usar_rango_edad": usar_rango_edad,
        "edad_min": edad_min_str,
        "edad_max": edad_max_str,
        # ═══════════════════════════════════════════════════════════════
        # NUEVOS FILTROS MINISTERIALES EN CONTEXTO
        # ═══════════════════════════════════════════════════════════════
        "rol_ministerial": rol_ministerial,
        "estado_ministerial": estado_ministerial,
        "tiene_credenciales_filtro": tiene_credenciales_filtro,
        "roles_ministeriales_choices": roles_ministeriales_choices,
        "estados_ministeriales_choices": estados_ministeriales_choices,
    }

    return miembros, filtros_context



# -------------------------------------
# LISTA DE MIEMBROS
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def miembro_lista(request):
    """
    Listado general de miembros (versión normal y versión imprimible).
    Usa la función filtrar_miembros para aplicar todos los filtros.
    IMPORTANTE: excluye a los nuevos creyentes (nuevo_creyente=True).
    """

    # Base: solo miembros que NO son nuevos creyentes
    miembros_base = Miembro.objects.filter(nuevo_creyente=False)

    miembros, filtros_context = filtrar_miembros(request, miembros_base)

    # Edad mínima oficial para mostrar en el texto de la plantilla
    edad_minima = get_edad_minima_miembro_oficial()

    context = {
        "miembros": miembros,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
        "modo_pdf": False,   # 👈 importante para distinguir vista normal de PDF
    }
    # Mezclamos con todos los filtros (para que el formulario recuerde el estado)
    context.update(filtros_context)

    return render(
        request,
        "miembros_app/reportes/listado_miembros.html",
        context,
    )

# -------------------------------------
# NOTIFICACIÓN POR CORREO: NUEVO MIEMBRO
# -------------------------------------
from core.models import ConfiguracionSistema
from core.utils_email import enviar_correo_sistema

def notificar_nuevo_miembro(miembro, request=None):
    """
    Envía un correo al correo oficial configurado cuando se registra un nuevo miembro.
    """

    config = ConfiguracionSistema.load()
    destinatario = config.email_oficial or settings.DEFAULT_FROM_EMAIL

    subject = f"Nuevo miembro: {miembro.nombres} {miembro.apellidos}"

    # Fecha de ingreso (evitar errores)
    if getattr(miembro, "fecha_ingreso_iglesia", None):
        fecha_ingreso = miembro.fecha_ingreso_iglesia.strftime("%d/%m/%Y")
    else:
        fecha_ingreso = "-"

    body_html = f"""
        <p>Hola,</p>

        <p>Se ha registrado un nuevo miembro en el sistema <strong>Soid_Tf_2</strong>:</p>

        <p>
            <strong>Nombre:</strong> {miembro.nombres} {miembro.apellidos}<br>
            <strong>Estado:</strong> {miembro.get_estado_miembro_display() or "Sin estado"}<br>
            <strong>Fecha de ingreso:</strong> {fecha_ingreso}
        </p>

        <p>Puedes consultar más detalles desde el sistema.</p>

        <p style="margin-top:16px;">
            Bendiciones,<br>
            <strong>Soid_Tf_2</strong>
        </p>
    """

    # Construir URL del detalle
    button_url = None
    if request is not None:
        try:
            detalle_url = reverse("miembros_app:editar", args=[miembro.pk])
            button_url = request.build_absolute_uri(detalle_url)
        except:
            button_url = None

    enviar_correo_sistema(
        subject=subject,
        heading="Nuevo miembro registrado",
        subheading="Un nuevo miembro ha sido añadido al sistema.",
        body_html=body_html,
        destinatarios=destinatario,
        button_url=button_url,
        button_text="Ver ficha del miembro" if button_url else None,
        meta_text="Correo generado por Soid_Tf_2 automáticamente.",
    )


# -------------------------------------
# CREAR MIEMBRO
# -------------------------------------
@login_required
@permission_required("miembros_app.add_miembro", raise_exception=True)
def miembro_crear(request):
    """
    Vista normal para crear un miembro.
    Con la lógica de edad mínima tomando el valor desde los parámetros del sistema.
    """

    edad_minima = get_edad_minima_miembro_oficial()

    if request.method == "POST":
        form = MiembroForm(request.POST, request.FILES)

        if form.is_valid():
            miembro = form.save(commit=False)

            # Fecha de ingreso automática
            if not miembro.fecha_ingreso_iglesia:
                miembro.fecha_ingreso_iglesia = date.today()
            # Calcular edad
            edad = None
            if hasattr(miembro, "calcular_edad"):
                edad = miembro.calcular_edad()
            elif miembro.fecha_nacimiento:
                hoy = date.today()
                fn = miembro.fecha_nacimiento
                edad = hoy.year - fn.year
                if (hoy.month, hoy.day) < (fn.month, fn.day):
                    edad -= 1

            # Lógica de edad mínima
            if edad is not None and edad < edad_minima:
                if miembro.estado_miembro:
                    miembro.estado_miembro = ""
                    messages.info(
                        request,
                        (
                            f"Este registro es menor de {edad_minima} años. "
                            "Se ha guardado sin estado de miembro."
                        ),
                    )

            miembro.save()
            
          

            # 🔔 Crear notificación del sistema
            try:
                # URL al detalle del miembro (ej. /miembros/miembro/123/)
                url_detalle = reverse("miembros_app:detalle", args=[miembro.pk])

                # OJO: crear_notificacion espera:
                # (usuario, titulo, mensaje="", url_name=None, tipo="info")
                crear_notificacion(
                    usuario=request.user,
                    titulo="Nuevo miembro registrado",
                    mensaje=f"{miembro.nombres} {miembro.apellidos} ha sido añadido al sistema.",
                    # Le pasamos la URL completa como url_name; si no es un 'name',
                    # la función la usará tal cual como destino.
                    url_name=url_detalle,
                    tipo="success",
                )
            except Exception as e:
                print("Error creando notificación:", e)
           

            # --- AQUÍ DECIDIMOS SEGÚN EL BOTÓN PULSADO ---
            if "guardar_y_nuevo" in request.POST:
                messages.success(
                    request,
                    "Miembro creado correctamente. Puedes registrar otro."
                )
                # Volvemos al formulario de creación limpio
                return redirect("miembros_app:crear")
            else:
                # Botón 'Guardar' normal
                messages.success(request, "Miembro creado correctamente.")
                return redirect("miembros_app:lista")
        else:
            # Si no es válido, se vuelve a mostrar el formulario con errores
            messages.error(
                request,
                "Hay errores en el formulario. Revisa los campos marcados en rojo."
            )
    else:
        form = MiembroForm()

    context = {
        "form": form,
        "modo": "crear",
        "miembro": None,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/miembro_form.html", context)

def _get_padres_ids(miembro_id):
    return set(
        MiembroRelacion.objects
        .filter(miembro_id=miembro_id, tipo_relacion__in=["padre", "madre"])
        .values_list("familiar_id", flat=True)
    )

def _get_hijos_ids(miembro_id):
    return set(
        MiembroRelacion.objects
        .filter(miembro_id=miembro_id, tipo_relacion="hijo")
        .values_list("familiar_id", flat=True)
    )

def _get_conyuge_ids(miembro_id):
    return set(
        MiembroRelacion.objects
        .filter(miembro_id=miembro_id, tipo_relacion="conyuge")
        .values_list("familiar_id", flat=True)
    )



"""
═══════════════════════════════════════════════════════════════════════════════
INFERENCIA INTELIGENTE DE RELACIONES FAMILIARES v4 - DEFINITIVA
═══════════════════════════════════════════════════════════════════════════════

CORREGIDO COMPLETAMENTE:
- Todas las inferencias ahora usan padres/hijos COMPLETOS (directos + inferidos)
- Funciona en cascada: si A es padre inferido de B, y B es padre de C, 
  entonces A es abuelo inferido de C

CASOS CUBIERTOS:
✅ Padres inferidos (cónyuge de mi padre/madre)
✅ Hijos inferidos (hijos de mi cónyuge)  
✅ Hermanos (compartimos padre/madre, incluyendo inferidos)
✅ Abuelos (padres de mis padres, incluyendo inferidos)
✅ Tíos (hermanos de mis padres, usando abuelos completos)
✅ Sobrinos (hijos de mis hermanos, incluyendo inferidos)
✅ Primos (hijos de mis tíos)
✅ Nietos (hijos de mis hijos, incluyendo inferidos)
✅ Cuñados (cónyuge de hermano + hermanos de cónyuge)
✅ Suegros (padres de mi cónyuge, incluyendo inferidos)
✅ Yernos/Nueras (cónyuges de mis hijos)
✅ Consuegros (padres de mis yernos/nueras)

═══════════════════════════════════════════════════════════════════════════════
"""

from django.db.models import Q


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════════

def _obtener_relaciones_directas(miembro_id, MiembroRelacion):
    """
    Obtiene todas las relaciones directas de un miembro, normalizadas.
    Retorna: dict con sets de IDs por tipo de relación.
    """
    padres = set()
    hijos = set()
    hermanos = set()
    conyuges = set()
    
    rels = (
        MiembroRelacion.objects
        .filter(Q(miembro_id=miembro_id) | Q(familiar_id=miembro_id))
        .select_related("miembro", "familiar")
    )
    
    for rel in rels:
        if rel.miembro_id == miembro_id:
            otro_id = rel.familiar_id
            tipo = rel.tipo_relacion
        else:
            otro_id = rel.miembro_id
            tipo = MiembroRelacion.inverse_tipo(rel.tipo_relacion, rel.miembro.genero)
        
        if tipo in ("padre", "madre"):
            padres.add(otro_id)
        elif tipo == "hijo":
            hijos.add(otro_id)
        elif tipo == "hermano":
            hermanos.add(otro_id)
        elif tipo == "conyuge":
            conyuges.add(otro_id)
    
    return {
        "padres": padres,
        "hijos": hijos,
        "hermanos": hermanos,
        "conyuges": conyuges,
    }


def _obtener_padres_completos(miembro_id, MiembroRelacion, _cache=None):
    """
    Obtiene TODOS los padres de un miembro (directos + inferidos por cónyuge).
    Retorna: (padres_directos, padres_inferidos)
    """
    if _cache is None:
        _cache = {}
    
    if miembro_id in _cache:
        return _cache[miembro_id]
    
    rels = _obtener_relaciones_directas(miembro_id, MiembroRelacion)
    padres_directos = rels["padres"]
    
    # Padres inferidos = cónyuges de los padres directos
    padres_inferidos = set()
    
    if padres_directos:
        rels_padres = (
            MiembroRelacion.objects
            .filter(
                Q(miembro_id__in=padres_directos, tipo_relacion="conyuge") |
                Q(familiar_id__in=padres_directos, tipo_relacion="conyuge")
            )
        )
        
        for rel in rels_padres:
            if rel.miembro_id in padres_directos:
                conyuge_id = rel.familiar_id
            else:
                conyuge_id = rel.miembro_id
            
            if conyuge_id not in padres_directos and conyuge_id != miembro_id:
                padres_inferidos.add(conyuge_id)
    
    resultado = (padres_directos, padres_inferidos)
    _cache[miembro_id] = resultado
    return resultado


def _obtener_hijos_completos(miembro_id, MiembroRelacion, _cache=None):
    """
    Obtiene TODOS los hijos de un miembro (directos + inferidos por cónyuge).
    Retorna: (hijos_directos, hijos_inferidos)
    """
    if _cache is None:
        _cache = {}
    
    cache_key = f"hijos_{miembro_id}"
    if cache_key in _cache:
        return _cache[cache_key]
    
    rels = _obtener_relaciones_directas(miembro_id, MiembroRelacion)
    hijos_directos = rels["hijos"]
    conyuges = rels["conyuges"]
    
    # Hijos inferidos = hijos de mis cónyuges que no son míos
    hijos_inferidos = set()
    
    if conyuges:
        for conyuge_id in conyuges:
            rels_conyuge = _obtener_relaciones_directas(conyuge_id, MiembroRelacion)
            hijos_conyuge = rels_conyuge["hijos"]
            
            for hijo_id in hijos_conyuge:
                if hijo_id not in hijos_directos and hijo_id != miembro_id:
                    hijos_inferidos.add(hijo_id)
    
    resultado = (hijos_directos, hijos_inferidos)
    _cache[cache_key] = resultado
    return resultado


def _obtener_hermanos_completos(miembro_id, todos_padres_ids, MiembroRelacion):
    """
    Obtiene todos los hermanos (personas que comparten al menos un padre).
    """
    hermanos = set()
    
    if todos_padres_ids:
        # Personas que dicen "X es mi padre/madre" donde X está en todos_padres_ids
        hermanos = set(
            MiembroRelacion.objects
            .filter(tipo_relacion__in=["padre", "madre"], familiar_id__in=todos_padres_ids)
            .exclude(miembro_id=miembro_id)
            .values_list("miembro_id", flat=True)
        )
        
        # Personas que mis padres dicen "X es mi hijo"
        hijos_de_padres = set(
            MiembroRelacion.objects
            .filter(miembro_id__in=todos_padres_ids, tipo_relacion="hijo")
            .exclude(familiar_id=miembro_id)
            .values_list("familiar_id", flat=True)
        )
        hermanos |= hijos_de_padres
    
    return hermanos


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════

def calcular_parentescos_inferidos(miembro):
    """
    Calcula TODOS los parentescos inferidos de un miembro.
    Usa inferencias en cascada para máxima precisión.
    """
    from .models import MiembroRelacion, Miembro
    
    mi_id = miembro.id
    cache = {}  # Cache para evitar queries repetidos

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 1: Obtener MIS relaciones directas
    # ═══════════════════════════════════════════════════════════════════════════
    mis_rels = _obtener_relaciones_directas(mi_id, MiembroRelacion)
    
    padres_directos = mis_rels["padres"]
    hijos_directos = mis_rels["hijos"]
    hermanos_directos = mis_rels["hermanos"]
    conyuges_directos = mis_rels["conyuges"]

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 2: MIS PADRES COMPLETOS (directos + inferidos)
    # ═══════════════════════════════════════════════════════════════════════════
    padres_dir, padres_inf = _obtener_padres_completos(mi_id, MiembroRelacion, cache)
    todos_mis_padres = padres_dir | padres_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 3: MIS HIJOS COMPLETOS (directos + inferidos)
    # ═══════════════════════════════════════════════════════════════════════════
    hijos_dir, hijos_inf = _obtener_hijos_completos(mi_id, MiembroRelacion, cache)
    todos_mis_hijos = hijos_dir | hijos_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 4: MIS HERMANOS (compartimos padre/madre)
    # ═══════════════════════════════════════════════════════════════════════════
    hermanos_inferidos = _obtener_hermanos_completos(mi_id, todos_mis_padres, MiembroRelacion)
    hermanos_inferidos |= hermanos_directos
    todos_mis_hermanos = hermanos_inferidos.copy()

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 5: MIS ABUELOS (padres de mis padres, INCLUYENDO INFERIDOS)
    # ═══════════════════════════════════════════════════════════════════════════
    abuelos_ids = set()
    
    for padre_id in todos_mis_padres:
        padres_de_padre_dir, padres_de_padre_inf = _obtener_padres_completos(
            padre_id, MiembroRelacion, cache
        )
        abuelos_ids |= padres_de_padre_dir
        abuelos_ids |= padres_de_padre_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 6: MIS TÍOS (hermanos de mis padres)
    # Hermano de mi padre = alguien que comparte padre con mi padre
    # ═══════════════════════════════════════════════════════════════════════════
    tios_ids = set()
    
    for padre_id in todos_mis_padres:
        # Obtener los padres de mi padre (mis abuelos por esa línea)
        padres_de_padre_dir, padres_de_padre_inf = _obtener_padres_completos(
            padre_id, MiembroRelacion, cache
        )
        abuelos_linea = padres_de_padre_dir | padres_de_padre_inf
        
        # Los hermanos de mi padre = hijos de mis abuelos que no son mi padre
        hermanos_del_padre = _obtener_hermanos_completos(padre_id, abuelos_linea, MiembroRelacion)
        tios_ids |= hermanos_del_padre

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 7: MIS SOBRINOS (hijos de mis hermanos, INCLUYENDO INFERIDOS)
    # ═══════════════════════════════════════════════════════════════════════════
    sobrinos_ids = set()
    
    for hermano_id in todos_mis_hermanos:
        hijos_hermano_dir, hijos_hermano_inf = _obtener_hijos_completos(
            hermano_id, MiembroRelacion, cache
        )
        sobrinos_ids |= hijos_hermano_dir
        sobrinos_ids |= hijos_hermano_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 8: MIS PRIMOS (hijos de mis tíos)
    # ═══════════════════════════════════════════════════════════════════════════
    primos_ids = set()
    
    for tio_id in tios_ids:
        hijos_tio_dir, hijos_tio_inf = _obtener_hijos_completos(
            tio_id, MiembroRelacion, cache
        )
        primos_ids |= hijos_tio_dir
        primos_ids |= hijos_tio_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 9: MIS NIETOS (hijos de mis hijos, INCLUYENDO INFERIDOS)
    # ═══════════════════════════════════════════════════════════════════════════
    nietos_ids = set()
    
    for hijo_id in todos_mis_hijos:
        hijos_hijo_dir, hijos_hijo_inf = _obtener_hijos_completos(
            hijo_id, MiembroRelacion, cache
        )
        nietos_ids |= hijos_hijo_dir
        nietos_ids |= hijos_hijo_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 10: MIS CUÑADOS
    # - Cónyuges de mis hermanos
    # - Hermanos de mi cónyuge
    # ═══════════════════════════════════════════════════════════════════════════
    cunados_ids = set()
    
    # Cónyuges de mis hermanos
    for hermano_id in todos_mis_hermanos:
        rels_hermano = _obtener_relaciones_directas(hermano_id, MiembroRelacion)
        cunados_ids |= rels_hermano["conyuges"]
    
    # Hermanos de mi cónyuge
    for conyuge_id in conyuges_directos:
        padres_conyuge_dir, padres_conyuge_inf = _obtener_padres_completos(
            conyuge_id, MiembroRelacion, cache
        )
        padres_conyuge = padres_conyuge_dir | padres_conyuge_inf
        
        hermanos_conyuge = _obtener_hermanos_completos(conyuge_id, padres_conyuge, MiembroRelacion)
        cunados_ids |= hermanos_conyuge

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 11: MIS SUEGROS (padres de mi cónyuge, INCLUYENDO INFERIDOS)
    # ═══════════════════════════════════════════════════════════════════════════
    suegros_ids = set()
    
    for conyuge_id in conyuges_directos:
        padres_conyuge_dir, padres_conyuge_inf = _obtener_padres_completos(
            conyuge_id, MiembroRelacion, cache
        )
        suegros_ids |= padres_conyuge_dir
        suegros_ids |= padres_conyuge_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 12: MIS YERNOS/NUERAS (cónyuges de mis hijos)
    # ═══════════════════════════════════════════════════════════════════════════
    yernos_ids = set()
    
    for hijo_id in todos_mis_hijos:
        rels_hijo = _obtener_relaciones_directas(hijo_id, MiembroRelacion)
        yernos_ids |= rels_hijo["conyuges"]

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 13: MIS CONSUEGROS (padres de mis yernos/nueras)
    # ═══════════════════════════════════════════════════════════════════════════
    consuegros_ids = set()
    
    for yerno_id in yernos_ids:
        padres_yerno_dir, padres_yerno_inf = _obtener_padres_completos(
            yerno_id, MiembroRelacion, cache
        )
        consuegros_ids |= padres_yerno_dir
        consuegros_ids |= padres_yerno_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 14: BISABUELOS (padres de mis abuelos)
    # ═══════════════════════════════════════════════════════════════════════════
    bisabuelos_ids = set()
    
    for abuelo_id in abuelos_ids:
        padres_abuelo_dir, padres_abuelo_inf = _obtener_padres_completos(
            abuelo_id, MiembroRelacion, cache
        )
        bisabuelos_ids |= padres_abuelo_dir
        bisabuelos_ids |= padres_abuelo_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # PASO 15: BISNIETOS (hijos de mis nietos)
    # ═══════════════════════════════════════════════════════════════════════════
    bisnietos_ids = set()
    
    for nieto_id in nietos_ids:
        hijos_nieto_dir, hijos_nieto_inf = _obtener_hijos_completos(
            nieto_id, MiembroRelacion, cache
        )
        bisnietos_ids |= hijos_nieto_dir
        bisnietos_ids |= hijos_nieto_inf

    # ═══════════════════════════════════════════════════════════════════════════
    # LIMPIEZA: Quitar duplicados y relaciones directas
    # ═══════════════════════════════════════════════════════════════════════════
    
    ids_directos = padres_directos | hijos_directos | hermanos_directos | conyuges_directos
    
    # Todos los sets de inferidos
    todos_sets = [
        padres_inf, hijos_inf, hermanos_inferidos, abuelos_ids, tios_ids,
        sobrinos_ids, primos_ids, nietos_ids, cunados_ids, suegros_ids,
        yernos_ids, consuegros_ids, bisabuelos_ids, bisnietos_ids
    ]
    
    for s in todos_sets:
        s.discard(mi_id)
        s -= ids_directos

    # Quitar hermanos directos del set de hermanos inferidos
    hermanos_inferidos -= hermanos_directos

    # ═══════════════════════════════════════════════════════════════════════════
    # CONSTRUIR RESULTADO
    # ═══════════════════════════════════════════════════════════════════════════
    
    ids_total = set()
    for s in todos_sets:
        ids_total |= s
    
    miembros_map = {
        m.id: m 
        for m in Miembro.objects.filter(id__in=ids_total).only("id", "nombres", "apellidos", "genero")
    }

    def pack(ids_set, tipo, razon=""):
        out = []
        for mid in ids_set:
            otro = miembros_map.get(mid)
            if not otro:
                continue
            out.append({
                "otro": otro,
                "tipo": tipo,
                "tipo_label": MiembroRelacion.label_por_genero(tipo, otro.genero),
                "inferido": True,
                "razon": razon,
            })
        return out

    inferidos = []
    
    inferidos += pack(padres_inf, "padre", "Cónyuge de tu padre/madre")
    inferidos += pack(hijos_inf, "hijo", "Hijo/a de tu cónyuge")
    inferidos += pack(hermanos_inferidos, "hermano", "Comparten padre/madre")
    inferidos += pack(abuelos_ids, "abuelo", "Padre/madre de tu padre/madre")
    inferidos += pack(bisabuelos_ids, "bisabuelo", "Padre/madre de tu abuelo/a")
    inferidos += pack(tios_ids, "tio", "Hermano/a de tu padre/madre")
    inferidos += pack(sobrinos_ids, "sobrino", "Hijo/a de tu hermano/a")
    inferidos += pack(primos_ids, "primo", "Hijo/a de tu tío/a")
    inferidos += pack(nietos_ids, "nieto", "Hijo/a de tu hijo/a")
    inferidos += pack(bisnietos_ids, "bisnieto", "Hijo/a de tu nieto/a")
    inferidos += pack(cunados_ids, "cunado", "Cónyuge de hermano/a o hermano/a de cónyuge")
    inferidos += pack(suegros_ids, "suegro", "Padre/madre de tu cónyuge")
    inferidos += pack(yernos_ids, "yerno", "Cónyuge de tu hijo/a")
    inferidos += pack(consuegros_ids, "consuegro", "Padre/madre del cónyuge de tu hijo/a")

    return inferidos
# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN ADICIONAL: Ver familia completa de un miembro
# ═══════════════════════════════════════════════════════════════════════════════

def obtener_familia_completa(miembro):
    """
    Devuelve TODAS las relaciones de un miembro:
    - Directas (guardadas en DB)
    - Inferidas (calculadas)
    
    Útil para mostrar todo en una sola lista.
    """
    from .models import MiembroRelacion
    
    mi_id = miembro.id
    
    # 1) Relaciones directas (normalizadas para mí)
    relaciones_directas = []
    
    rels_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro_id=mi_id) | Q(familiar_id=mi_id))
        .select_related("miembro", "familiar")
    )
    
    for rel in rels_qs:
        if rel.miembro_id == mi_id:
            otro = rel.familiar
            tipo = rel.tipo_relacion
        else:
            otro = rel.miembro
            tipo = MiembroRelacion.inverse_tipo(rel.tipo_relacion, otro.genero)
        
        relaciones_directas.append({
            "otro": otro,
            "tipo": tipo,
            "tipo_label": MiembroRelacion.label_por_genero(tipo, otro.genero),
            "inferido": False,
            "vive_junto": rel.vive_junto,
            "es_responsable": rel.es_responsable,
            "notas": rel.notas,
        })
    
    # 2) Relaciones inferidas
    relaciones_inferidas = calcular_parentescos_inferidos(miembro)
    
    # 3) Combinar (sin duplicados)
    ids_directos = {r["otro"].id for r in relaciones_directas}
    
    familia_completa = relaciones_directas.copy()
    
    for rel in relaciones_inferidas:
        if rel["otro"].id not in ids_directos:
            familia_completa.append(rel)
    
    return familia_completa


# -------------------------------------
# EDITAR MIEMBRO
# -------------------------------------
class MiembroUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = "miembros_app.change_miembro"
    raise_exception = True

    
    def get(self, request, pk):
        miembro = get_object_or_404(Miembro, pk=pk)
        form = MiembroForm(instance=miembro)

        edad_minima = get_edad_minima_miembro_oficial()

       

        familiares_qs = (
            MiembroRelacion.objects
            .filter(miembro=miembro)
            .select_related("familiar")
        )

        familiares_ids = familiares_qs.values_list("familiar_id", flat=True)

        todos_miembros = (
            Miembro.objects
            .exclude(pk=miembro.pk)
            .exclude(pk__in=familiares_ids)
            .order_by("nombres", "apellidos")
        )

        bloquear_identidad = _miembro_tiene_asignacion_en_unidades(miembro)

        bloquear_estado = _miembro_tiene_asignacion_en_unidades(miembro)
        rel_form = MiembroRelacionForm()
        
        relaciones_inferidas = calcular_parentescos_inferidos(miembro)

        context = {
            "form": form,
            "miembro": miembro,
            "rel_form": rel_form,  # ✅ NUEVO
            "modo": "editar",
            "todos_miembros": todos_miembros,
            "familiares": familiares_qs,
            "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
              "bloquear_estado": bloquear_estado,
              "bloquear_identidad": bloquear_identidad,
               "TIPOS_RELACION_CHOICES": MiembroRelacion.TIPO_RELACION_CHOICES,
               "relaciones_inferidas": relaciones_inferidas,



            
        }
        return render(request, "miembros_app/miembro_form.html", context)

    def post(self, request, pk):
        miembro = get_object_or_404(Miembro, pk=pk)
        edad_minima = get_edad_minima_miembro_oficial()

        salida_antes = (not miembro.activo and miembro.fecha_salida is not None)

        # --- SI VIENE DEL BOTÓN "AGREGAR FAMILIAR" ---
        if "agregar_familiar" in request.POST:
            rel_form = MiembroRelacionForm(
                request.POST,
                miembro=miembro
            )
            if rel_form.is_valid():
                relacion = rel_form.save(commit=False)
                relacion.miembro = miembro
                relacion.save()
                sync_familia_inteligente_por_relacion(relacion)
                messages.success(request, "Familiar agregado correctamente.")
            else:
                for field, errs in rel_form.errors.items():
                    for e in errs:
                        messages.error(request, f"{field}: {e}")

            return redirect(f"{request.path}?tab=familiares")

        # --- GUARDADO NORMAL DEL MIEMBRO ---
        form = MiembroForm(request.POST, request.FILES, instance=miembro)

        if form.is_valid():
            miembro_editado = form.save(commit=False)

            # =============================
            # BLOQUEO: NO CAMBIAR GÉNERO / FECHA NACIMIENTO SI ESTÁ EN UNIDADES
            # =============================
            genero_antes = miembro.genero
            fn_antes = miembro.fecha_nacimiento

            genero_despues = miembro_editado.genero
            fn_despues = miembro_editado.fecha_nacimiento

            cambio_genero = (genero_antes != genero_despues)
            cambio_fn = (fn_antes != fn_despues)

            if (cambio_genero or cambio_fn) and _miembro_tiene_asignacion_en_unidades(miembro):
                # Volvemos a renderizar la misma pantalla con el formulario (NO guarda)
                familiares_qs = (
                    MiembroRelacion.objects
                    .filter(miembro=miembro)
                    .select_related("familiar")
                )
                familiares_ids = familiares_qs.values_list("familiar_id", flat=True)

                todos_miembros = (
                    Miembro.objects
                    .exclude(pk=miembro.pk)
                    .exclude(pk__in=familiares_ids)
                    .order_by("nombres", "apellidos")
                )

                context = {
                    "form": form,
                    "miembro": miembro,
                    "modo": "editar",
                    "todos_miembros": todos_miembros,
                    "familiares": familiares_qs,
                    "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
                    "bloquear_identidad": True,
                    "TIPOS_RELACION_CHOICES": MiembroRelacion.TIPO_RELACION_CHOICES,
                }
                messages.error(
                    request,
                    "Acción bloqueada: no puedes cambiar género o fecha de nacimiento "
                    "mientras el miembro esté asignado a una o más unidades."
                )
                return render(request, "miembros_app/miembro_form.html", context)

            # =============================
            # BLOQUEO: NO CAMBIAR ESTADO SI ESTÁ EN UNIDADES
            # (si tienes este bloque en tu código, mantenlo igual; aquí asumo que ya existe arriba o abajo)
            # =============================

            # Guardar miembro
            miembro_editado.save()

            # =============================
            # ✅ SINCRONIZAR FAMILIARES (CREAR / ACTUALIZAR / ELIMINAR)
            # =============================
            ids = request.POST.getlist("familiares_miembro_id[]")
            tipos = request.POST.getlist("familiares_tipo_relacion[]")
            vive_list = request.POST.getlist("familiares_vive_junto[]")
            resp_list = request.POST.getlist("familiares_es_responsable[]")
            notas_list = request.POST.getlist("familiares_notas[]")

            # Normalizamos longitudes (si no viene nada, interpretamos como "quedó vacío")
            if not (len(ids) == len(tipos) == len(vive_list) == len(resp_list) == len(notas_list)):
                messages.error(request, "No se pudieron procesar los familiares: datos incompletos.")
            else:
                # 1) Lo que el usuario dejó en pantalla (POST)
                posted_ids = []
                payload = {}  # familiar_id -> dict campos

                for i in range(len(ids)):
                    try:
                        familiar_id = int(ids[i])
                    except (TypeError, ValueError):
                        continue

                    # No permitimos relacionarse consigo mismo
                    if familiar_id == miembro_editado.pk:
                        messages.error(
                            request,
                            "No puedes asignar un miembro como familiar de sí mismo."
                        )
                        continue

                    tipo = (tipos[i] or "otro").strip()

                    payload[familiar_id] = {
                        "tipo_relacion": tipo,
                        "vive_junto": (vive_list[i] == "1"),
                        "es_responsable": (resp_list[i] == "1"),
                        "notas": (notas_list[i] or "").strip(),
                    }
                    posted_ids.append(familiar_id)

                posted_set = set(posted_ids)

                # 2) Lo que existe en BD actualmente
                existentes_qs = MiembroRelacion.objects.filter(miembro=miembro_editado)
                existentes_set = set(existentes_qs.values_list("familiar_id", flat=True))

                # 3) ELIMINAR los que ya no están
                a_borrar = existentes_set - posted_set
                if a_borrar:
                    # Borra relación directa
                    MiembroRelacion.objects.filter(miembro=miembro_editado, familiar_id__in=a_borrar).delete()
                    # Borra inversa
                    MiembroRelacion.objects.filter(miembro_id__in=a_borrar, familiar=miembro_editado).delete()

                # 4) CREAR o ACTUALIZAR los que están
                for familiar_id, data in payload.items():
                    rel, _created = MiembroRelacion.objects.update_or_create(
                        miembro=miembro_editado,
                        familiar_id=familiar_id,
                        defaults=data
                    )

                    # Inversa automática: describe a "miembro_editado" desde el punto de vista del familiar
                    tipo_inverso = MiembroRelacion.inverse_tipo(
                        data["tipo_relacion"],
                        genero_persona_invertida=miembro_editado.genero
                    )

                    MiembroRelacion.objects.update_or_create(
                        miembro_id=familiar_id,
                        familiar=miembro_editado,
                        defaults={
                            "tipo_relacion": tipo_inverso,
                            "vive_junto": rel.vive_junto,
                            "es_responsable": False,
                            "notas": "",
                        },
                    )

            # Notificación salida (tu lógica)
            salida_despues = (not miembro_editado.activo and miembro_editado.fecha_salida is not None)
            if (not salida_antes) and salida_despues:
                try:
                    crear_notificacion(
                        request.user,
                        titulo="Miembro dado de salida",
                        mensaje=f"Se ha dado salida al miembro {miembro_editado.nombres} {miembro_editado.apellidos}.",
                        url_name="miembros_app:detalle",
                        kwargs={"pk": miembro_editado.pk},
                        tipo="warning",
                    )
                except Exception:
                    pass

            messages.success(request, "Miembro actualizado correctamente.")
            return redirect("miembros_app:lista")

        # Si hay errores, recargar pantalla
        familiares_qs = (
            MiembroRelacion.objects
            .filter(miembro=miembro)
            .select_related("familiar")
        )
        familiares_ids = familiares_qs.values_list("familiar_id", flat=True)

        todos_miembros = (
            Miembro.objects
            .exclude(pk=miembro.pk)
            .exclude(pk__in=familiares_ids)
            .order_by("nombres", "apellidos")
        )

        context = {
            "form": form,
            "miembro": miembro,
            "modo": "editar",
            "todos_miembros": todos_miembros,
            "familiares": familiares_qs,
            "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
            "TIPOS_RELACION_CHOICES": MiembroRelacion.TIPO_RELACION_CHOICES,
        }
        return render(request, "miembros_app/miembro_form.html", context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        miembro = self.get_object()
        context["hogar_principal"] = miembro.hogar_principal
        context["miembros_hogar"] = miembro.miembros_de_mi_hogar
        context["clan_familiar"] = miembro.clan_familiar
        context["hogares_clan"] = miembro.hogares_de_mi_clan

        return context




# -------------------------------------
# DETALLE DEL MIEMBRO
# -------------------------------------

class MiembroDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = "miembros_app.view_miembro"
    raise_exception = True

    def get(self, request, pk):
        miembro = get_object_or_404(Miembro, pk=pk)

        # =========================
        # FINANZAS
        # =========================
        movimientos_financieros = (
            MovimientoFinanciero.objects
            .filter(persona_asociada=miembro)
            .exclude(estado="anulado")
            .order_by("-fecha", "-creado_en")
        )

        total_aportes = (
            movimientos_financieros
            .filter(tipo="ingreso")
            .aggregate(total=Sum("monto"))["total"] or 0
        )

        # =========================
        # PERMISOS
        # =========================
        can_dar_salida = (
            request.user.is_authenticated
            and request.user.has_perm("miembros_app.change_miembro")
        )

        # =========================
        # FAMILIA (organizada en 4 categorías)
        # =========================
        relaciones_organizadas = obtener_relaciones_organizadas(miembro)
        familia_nuclear = relaciones_organizadas["familia_nuclear"]
        familia_origen = relaciones_organizadas["familia_origen"]
        familia_extendida = relaciones_organizadas["familia_extendida"]
        familia_politica = relaciones_organizadas["familia_politica"]

        edad_minima = get_edad_minima_miembro_oficial()

        # =========================
        # CONTEXT BASE
        # =========================
        context = {
            "miembro": miembro,
            # Familia organizada en 4 categorías
            "familia_nuclear": familia_nuclear,
            "familia_origen": familia_origen,
            "familia_extendida": familia_extendida,
            "familia_politica": familia_politica,
            # Resto del contexto
            "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
            "movimientos_financieros": movimientos_financieros,
            "total_aportes": total_aportes,
            "can_dar_salida": can_dar_salida,
            "unidades_resumen": [],
            "unidades_total": 0,
        }


                # =========================
        # 🔐 BLOQUEO PESTAÑA PRIVADA
        # =========================
        context["privado_desbloqueado"] = request.session.get(
            f"miembro_privado_{pk}", False
        )

        # =========================
        # 🔐 BLOQUEO HISTORIAL FINANCIERO
        # =========================
        context["finanzas_desbloqueado"] = request.session.get(
            f"miembro_finanzas_{pk}", False
        )


        # =========================
        # UNIDADES (TABLA: Unidad / Rol / Tipo)
        # =========================
        estructura_activa = _modulo_estructura_activo()

        if estructura_activa:
            UnidadCargo = _safe_get_model("estructura_app", "UnidadCargo")
            UnidadMembresia = _safe_get_model("estructura_app", "UnidadMembresia")

            if UnidadCargo and UnidadMembresia:
                cargos_qs = (
                    UnidadCargo.objects
                    .filter(miembo_fk=miembro, vigente=True, unidad__activa=True)
                    .select_related("unidad", "rol")
                )

                membresias_qs = (
                    UnidadMembresia.objects
                    .filter(miembo_fk=miembro, activo=True, unidad__activa=True)
                    .select_related("unidad", "rol")
                )

                resumen = []
                vistos = set()

                for c in cargos_qs:
                    key = ("CARGO", c.unidad_id, c.rol_id)
                    if key in vistos:
                        continue
                    vistos.add(key)

                    resumen.append({
                        "unidad": c.unidad.nombre if c.unidad_id else "—",
                        "rol": c.rol.nombre if c.rol_id else "Miembro",
                        "tipo": "Liderazgo",
                    })

                for m in membresias_qs:
                    key = ("MEMB", m.unidad_id, m.rol_id)
                    if key in vistos:
                        continue
                    vistos.add(key)

                    rol_tipo = (m.rol.tipo or "").upper() if m.rol else ""
                    tipo = "Trabajo" if rol_tipo == "TRABAJO" else "Participación"

                    resumen.append({
                        "unidad": m.unidad.nombre if m.unidad_id else "—",
                        "rol": m.rol.nombre if m.rol_id else "Miembro",
                        "tipo": tipo,
                    })

                orden_tipo = {"Liderazgo": 0, "Participación": 1, "Trabajo": 2}
                resumen.sort(key=lambda x: (orden_tipo.get(x.get("tipo"), 99), x["unidad"], x["rol"]))

                context["unidades_resumen"] = resumen
                context["unidades_total"] = len(resumen)

        return render(request, "miembros_app/miembros_detalle.html", context)


# -------------------------------------
# AGREGAR FAMILIAR (VERSIÓN ANTIGUA)
# -------------------------------------
@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def agregar_familiar(request, pk):
    miembro = get_object_or_404(Miembro, pk=pk)

    if request.method == "POST":
        form = MiembroRelacionForm(request.POST)
        if form.is_valid():
            relacion = form.save(commit=False)
            relacion.miembro = miembro
            relacion.save()
            messages.success(request, "Familiar agregado correctamente.")
        else:
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"{field}: {e}")

    return redirect("miembros_app:editar", pk=miembro.pk)



@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_enviar_ficha_email(request, pk):
    miembro = get_object_or_404(Miembro, pk=pk)
    config = ConfiguracionSistema.load()

    # Nombre que verá el destinatario en el adjunto
    nombre_adjunto_auto = f"ficha_miembro_{miembro.pk}.pdf"

    # Datos iniciales para el formulario
    destinatario_inicial = (
        miembro.email
        or config.email_oficial
        or settings.DEFAULT_FROM_EMAIL
    )
    asunto_inicial = f"Ficha del miembro: {miembro.nombres} {miembro.apellidos}"
    mensaje_inicial = (
        f"Le enviamos la ficha del miembro {miembro.nombres} {miembro.apellidos} "
        f"de la iglesia {config.nombre_iglesia or 'Torre Fuerte'}."
    )

     # 👇 AQUÍ EL CAMBIO IMPORTANTE: usar 'ficha' como en la plantilla
    ficha_url = request.build_absolute_uri(
        reverse("miembros_app:ficha", args=[miembro.pk])
    )

    if request.method == "POST":
        # Ya no necesitamos request.FILES porque el PDF lo generamos nosotros
        form = EnviarFichaMiembroEmailForm(request.POST)
        if form.is_valid():
            destinatario = form.cleaned_data["destinatario"]
            asunto = form.cleaned_data["asunto"] or asunto_inicial
            mensaje = form.cleaned_data["mensaje"] or mensaje_inicial

            try:
                # 1) Generar el PDF usando Chrome Headless desde la URL real
                pdf_bytes = generar_pdf_desde_url(ficha_url)

                # 2) Cuerpo HTML del correo
                body_html = (
                    f"<p>{mensaje}</p>"
                    "<p style='margin-top:16px;'>"
                    "Bendiciones,<br>"
                    "<strong>Soid_Tf_2</strong>"
                    "</p>"
                )

                # 3) Enviar correo con el helper genérico
                enviar_correo_sistema(
                    subject=asunto,
                    heading="Ficha de miembro",
                    subheading=f"{miembro.nombres} {miembro.apellidos}",
                    body_html=body_html,
                    destinatarios=destinatario,
                    meta_text="Correo enviado desde el sistema Soid_Tf_2.",
                    extra_context={"CFG": config},
                    # Adjuntamos el PDF EN MEMORIA, igual que en el listado
                    adjuntos=[(nombre_adjunto_auto, pdf_bytes)],
                )

                messages.success(
                    request,
                    f"Correo enviado correctamente a {destinatario}."
                )
                return redirect("miembros_app:detalle", pk=miembro.pk)

            except Exception as e:
                messages.error(
                    request,
                    f"No se pudo enviar el correo: {e}"
                )

    else:
        form = EnviarFichaMiembroEmailForm(
            initial={
                "destinatario": destinatario_inicial,
                "asunto": asunto_inicial,
                "mensaje": mensaje_inicial,
            }
        )

    context = {
        "form": form,
        "miembro": miembro,
        "titulo_pagina": "Enviar ficha por correo",
        "descripcion": "Completa los datos para enviar esta ficha por correo electrónico.",
        "objeto_label": f"Miembro: {miembro.nombres} {miembro.apellidos}",
        "url_cancelar": reverse("miembros_app:detalle", args=[miembro.pk]),
        "adjunto_auto_nombre": nombre_adjunto_auto,
    }
    return render(request, "core/enviar_email.html", context)

# -------------------------------------
# ELIMINAR FAMILIAR
# -------------------------------------
@login_required
@permission_required("miembros_app.delete_miembrorelacion", raise_exception=True)
@require_POST
def eliminar_familiar(request, relacion_id):
    relacion = get_object_or_404(MiembroRelacion, pk=relacion_id)
    miembro_pk = relacion.miembro.pk

    relacion.delete()
    messages.success(request, "Familiar quitado correctamente.")

    url = reverse("miembros_app:editar", kwargs={"pk": miembro_pk})
    return redirect(f"{url}?tab=familiares")


@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def miembro_ficha(request, pk):
    """
    Ficha pastoral imprimible para un miembro concreto.
    """
    miembro = get_object_or_404(Miembro, pk=pk)

    # Traer familiares igual que en el detalle
    relaciones_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro=miembro) | Q(familiar=miembro))
        .select_related("miembro", "familiar")
    )

    relaciones_familia = []
    parejas_vistas = set()

    for rel in relaciones_qs:
        if rel.tipo_relacion == "conyuge":
            pareja = frozenset({rel.miembro_id, rel.familiar_id})
            if pareja in parejas_vistas:
                continue
            parejas_vistas.add(pareja)
            relaciones_familia.append(rel)
        else:
            relaciones_familia.append(rel)

    edad_minima = get_edad_minima_miembro_oficial()

    context = {
        "miembro": miembro,
        "relaciones_familia": relaciones_familia,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/reportes/miembro_ficha.html", context)


# -------------------------------------
# REPORTES - PANTALLA PRINCIPAL
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reportes_miembros(request):
    """
    Pantalla principal de reportes del módulo de miembros.
    Solo muestra enlaces a los distintos reportes disponibles.
    """
    # Podríamos pasar contadores en el futuro; por ahora es estático.
    return render(request, "miembros_app/reportes/reportes_home.html", {})


# -------------------------------------
# REPORTE: LISTADO GENERAL IMPRIMIBLE
# -------------------------------------


@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_listado_miembros(request):
    """
    Vista de reporte imprimible profesional con header institucional.
    Aplica los mismos filtros que miembro_lista usando filtrar_miembros().
    """
    from core.models import ConfiguracionSistema
    
    # Obtener configuración institucional para el header
    CFG = ConfiguracionSistema.load()
    
    # =========================================================================
    # CRÍTICO: Usar la misma base que miembro_lista (excluir nuevos creyentes)
    # =========================================================================
    miembros_base = Miembro.objects.filter(nuevo_creyente=False)
    
    # Usar la función compartida para aplicar TODOS los filtros consistentemente
    miembros, filtros_context = filtrar_miembros(request, miembros_base)
    
    # Contar activos e inactivos del resultado filtrado
    if isinstance(miembros, list):
        # Si filtrar_miembros devolvió una lista (por filtro de rango de edad)
        activos_count = sum(1 for m in miembros if m.activo)
        inactivos_count = sum(1 for m in miembros if not m.activo)
    else:
        # Si es un queryset
        activos_count = miembros.filter(activo=True).count()
        inactivos_count = miembros.filter(activo=False).count()
    
    # Obtener los valores crudos de los filtros para los enlaces
    estado = request.GET.get("estado", "").strip()
    genero_filtro = request.GET.get("genero", "").strip()
    categoria_edad_filtro = request.GET.get("categoria_edad", "").strip()
    bautizado = request.GET.get("bautizado", "").strip()
    
    # Obtener labels para mostrar en metadatos del reporte
    estado_label = ""
    if estado:
        estado_dict = dict(Miembro.ESTADOS_MIEMBRO)
        estado_label = estado_dict.get(estado, estado)
    
    genero_label = ""
    if genero_filtro:
        genero_dict = dict(Miembro.GENERO_CHOICES)
        genero_label = genero_dict.get(genero_filtro, genero_filtro)
    
    categoria_label = ""
    if categoria_edad_filtro:
        categoria_dict = dict(Miembro.CATEGORIA_EDAD_CHOICES)
        categoria_label = categoria_dict.get(categoria_edad_filtro, categoria_edad_filtro)
    
    context = {
        "miembros": miembros,
        "query": filtros_context.get("query", ""),
        # Labels para mostrar en el reporte (texto legible)
        "estado": estado_label,
        "genero_filtro": genero_label,
        "categoria_edad_filtro": categoria_label,
        "bautizado": bautizado,
        # Conteos
        "activos_count": activos_count,
        "inactivos_count": inactivos_count,
        "CFG": CFG,
    }
    
    return render(request, "miembros_app/reportes/reporte_listado_miembros.html", context)

# -------------------------------------
# REPORTE: FICHA PASTORAL DEL MIEMBRO
# (segunda definición simplificada)
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def miembro_ficha(request, pk):
    """
    Ficha pastoral imprimible para un miembro concreto.
    """
    miembro = get_object_or_404(Miembro, pk=pk)

    # Traer familiares igual que en el detalle
    relaciones_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro=miembro) | Q(familiar=miembro))
        .select_related("miembro", "familiar")
    )

    relaciones_familia = []
    parejas_vistas = set()

    for rel in relaciones_qs:
        if rel.tipo_relacion == "conyuge":
            pareja = frozenset({rel.miembro_id, rel.familiar_id})
            if pareja in parejas_vistas:
                continue
            parejas_vistas.add(pareja)
            relaciones_familia.append(rel)
        else:
            relaciones_familia.append(rel)

    edad_minima = get_edad_minima_miembro_oficial()

    context = {
        "miembro": miembro,
        "relaciones_familia": relaciones_familia,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/reportes/miembro_ficha.html", context)


# --------------------------------------------
# REPORTE: MIEMBROS QUE SE FUERON / TRASLADOS
# --------------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_miembros_salida(request):
    """
    Reporte de miembros inactivos con filtros por fecha y razón de salida.
    Usa el campo razon_salida como ForeignKey a RazonSalidaMiembro.
    """

    query = request.GET.get("q", "").strip()
    fecha_desde_str = request.GET.get("fecha_desde", "").strip()
    fecha_hasta_str = request.GET.get("fecha_hasta", "").strip()
    razon_salida_id_str = request.GET.get("razon_salida", "").strip()

    # Base: solo miembros inactivos
    miembros = Miembro.objects.filter(activo=False)

    # Filtro de búsqueda general
    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query)
            | Q(apellidos__icontains=query)
            | Q(email__icontains=query)
            | Q(telefono__icontains=query)
            | Q(telefono_secundario__icontains=query)
        )

    # Filtros de fecha
    if fecha_desde_str:
        try:
            fecha_desde = date.fromisoformat(fecha_desde_str)
            miembros = miembros.filter(fecha_salida__gte=fecha_desde)
        except ValueError:
            pass

    if fecha_hasta_str:
        try:
            fecha_hasta = date.fromisoformat(fecha_hasta_str)
            miembros = miembros.filter(fecha_salida__lte=fecha_hasta)
        except ValueError:
            pass

    # Filtro por razón de salida (ForeignKey)
    razon_salida_id = None
    razon_salida_obj = None
    if razon_salida_id_str and razon_salida_id_str.isdigit():
        razon_salida_id = int(razon_salida_id_str)
        miembros = miembros.filter(razon_salida_id=razon_salida_id)
        razon_salida_obj = RazonSalidaMiembro.objects.filter(pk=razon_salida_id).first()

    # Razones disponibles (solo activas)
    razones_disponibles = RazonSalidaMiembro.objects.filter(activo=True).order_by(
        "orden", "nombre"
    )

    # Orden final
    miembros = miembros.order_by("-fecha_salida", "apellidos", "nombres")

    context = {
        "miembros": miembros,
        "query": query,
        "fecha_desde": fecha_desde_str,
        "fecha_hasta": fecha_hasta_str,
        "razones_disponibles": razones_disponibles,
        "razon_salida_id": razon_salida_id,
        "razon_salida_obj": razon_salida_obj,
    }

    return render(
        request,
        "miembros_app/reportes/reporte_miembros_salida.html",
        context,
    )


from django.shortcuts import render
from django.db.models import Q
from django.utils import timezone
from collections import defaultdict

from .models import Miembro, MiembroRelacion
from core.utils_config import get_config



"""
═══════════════════════════════════════════════════════════════════════════════
REPORTE DE RELACIONES FAMILIARES - VERSIÓN 2
═══════════════════════════════════════════════════════════════════════════════

CAMBIO PRINCIPAL:
- Las familias nucleares se muestran SIEMPRE por separado
- Si una familia nuclear es parte de una extendida, aparece en AMBAS secciones
- Ejemplo: Leonel + Victoria (padres) + Jhon + Esposa + Hijo
  → En Extendidas: toda la familia junta
  → En Nucleares: "Familia de Jhon" (Jhon + Esposa + Hijo) por separado
"""

from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from collections import defaultdict

# Ajusta estos imports según tu proyecto:
# from .models import Miembro, MiembroRelacion
# from core.utils_config import get_config


@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_relaciones_familiares(request):
    """
    Reporte: Familias de la Iglesia
    
    - Familias Extendidas: Grupos con múltiples núcleos conectados
    - Familias Nucleares: TODAS las unidades padres+hijos (incluso si son parte de extendida)
    - Parejas: Cónyuges sin hijos registrados
    """
    
    hoy = timezone.localdate()
    CFG = get_config()
    
    # ═══════════════════════════════════════════════════════════════════════════
    # PARÁMETROS DE FILTRO
    # ═══════════════════════════════════════════════════════════════════════════
    query = request.GET.get("q", "").strip()
    tipo_filtro = request.GET.get("tipo", "").strip()
    
    # ═══════════════════════════════════════════════════════════════════════════
    # OBTENER TODAS LAS RELACIONES
    # ═══════════════════════════════════════════════════════════════════════════
    relaciones_qs = (
        MiembroRelacion.objects
        .select_related("miembro", "familiar")
        .filter(miembro__activo=True, familiar__activo=True)
    )
    
    if query:
        relaciones_qs = relaciones_qs.filter(
            Q(miembro__nombres__icontains=query) |
            Q(miembro__apellidos__icontains=query) |
            Q(familiar__nombres__icontains=query) |
            Q(familiar__apellidos__icontains=query)
        )
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CONSTRUIR GRAFO DE CONEXIONES
    # ═══════════════════════════════════════════════════════════════════════════
    conexiones = defaultdict(set)
    relaciones_info = {}
    
    for rel in relaciones_qs:
        mid = rel.miembro_id
        fid = rel.familiar_id
        
        if mid == fid:
            continue
        
        conexiones[mid].add(fid)
        conexiones[fid].add(mid)
        
        # Guardar info de la relación
        relaciones_info[(mid, fid)] = {
            "tipo": rel.tipo_relacion,
            "vive_junto": rel.vive_junto,
            "es_responsable": rel.es_responsable,
        }
        
        # Guardar inversa
        tipo_inverso = MiembroRelacion.inverse_tipo(
            rel.tipo_relacion,
            genero_persona_invertida=rel.miembro.genero
        )
        relaciones_info[(fid, mid)] = {
            "tipo": tipo_inverso,
            "vive_junto": rel.vive_junto,
            "es_responsable": False,
        }
    
    # ═══════════════════════════════════════════════════════════════════════════
    # ENCONTRAR GRUPOS FAMILIARES (componentes conectados)
    # ═══════════════════════════════════════════════════════════════════════════
    visitados = set()
    grupos = []
    
    for persona_id in list(conexiones.keys()):
        if persona_id in visitados:
            continue
        
        grupo = set()
        cola = [persona_id]
        
        while cola:
            actual = cola.pop(0)
            if actual in visitados:
                continue
            visitados.add(actual)
            grupo.add(actual)
            
            for conectado in conexiones[actual]:
                if conectado not in visitados:
                    cola.append(conectado)
        
        if len(grupo) >= 2:
            grupos.append(grupo)
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CARGAR DATOS DE MIEMBROS
    # ═══════════════════════════════════════════════════════════════════════════
    todos_ids = set()
    for grupo in grupos:
        todos_ids.update(grupo)
    
    miembros_map = {}
    if todos_ids:
        miembros_map = {m.id: m for m in Miembro.objects.filter(id__in=todos_ids)}
    
    # ═══════════════════════════════════════════════════════════════════════════
    # FUNCIONES AUXILIARES
    # ═══════════════════════════════════════════════════════════════════════════
    
    def rel_tipo(a, b):
        """Tipo de relación desde a -> b."""
        info = relaciones_info.get((a, b))
        return info["tipo"] if info else None
    
    def son_conyuges(a, b):
        return rel_tipo(a, b) == "conyuge" or rel_tipo(b, a) == "conyuge"
    
    def es_hijo_de(hijo_id, padre_id):
        """Retorna True si hijo_id es hijo de padre_id."""
        return (
            rel_tipo(padre_id, hijo_id) == "hijo"
            or rel_tipo(hijo_id, padre_id) in ("padre", "madre")
        )
    
    def obtener_rol_genero(tipo, genero):
        es_femenino = (genero or "").lower() in ("femenino", "f", "mujer")
        roles = {
            "padre": "Padre",
            "madre": "Madre",
            "hijo": "Hija" if es_femenino else "Hijo",
            "conyuge": "Esposa" if es_femenino else "Esposo",
            "hermano": "Hermana" if es_femenino else "Hermano",
            "abuelo": "Abuela" if es_femenino else "Abuelo",
            "nieto": "Nieta" if es_femenino else "Nieto",
            "bisabuelo": "Bisabuela" if es_femenino else "Bisabuelo",
            "bisnieto": "Bisnieta" if es_femenino else "Bisnieto",
            "tio": "Tía" if es_femenino else "Tío",
            "sobrino": "Sobrina" if es_femenino else "Sobrino",
            "primo": "Prima" if es_femenino else "Primo",
            "suegro": "Suegra" if es_femenino else "Suegro",
            "cunado": "Cuñada" if es_femenino else "Cuñado",
            "yerno": "Nuera" if es_femenino else "Yerno",
            "consuegro": "Consuegra" if es_femenino else "Consuegro",
        }
        return roles.get(tipo, "Familiar")
    
    def es_cabeza_familia(miembro_id, grupo_ids):
        for otro_id in grupo_ids:
            if otro_id == miembro_id:
                continue
            key = (otro_id, miembro_id)
            if key in relaciones_info and relaciones_info[key].get("es_responsable"):
                return True
        return False
    
    def obtener_parejas_en_grupo(grupo_ids):
        """Encuentra todas las parejas (cónyuges) en un grupo."""
        parejas = []
        vistos = set()
        
        for id1 in grupo_ids:
            for id2 in grupo_ids:
                if id1 >= id2:
                    continue
                if son_conyuges(id1, id2):
                    if (id1, id2) not in vistos:
                        parejas.append((id1, id2))
                        vistos.add((id1, id2))
        
        return parejas
    
    def obtener_hijos_directos(padre_ids, todos_ids):
        """Obtiene los hijos directos de un conjunto de padres."""
        hijos = set()
        for padre_id in padre_ids:
            for otro_id in todos_ids:
                if otro_id in padre_ids:
                    continue
                if es_hijo_de(otro_id, padre_id):
                    hijos.add(otro_id)
        return hijos
    
    def construir_datos_persona(m, rol, es_cabeza=False):
        """Crea el dict de datos de una persona."""
        return {
            "id": m.id,
            "nombre_completo": f"{m.nombres} {m.apellidos}",
            "rol": rol,
            "edad": m.calcular_edad(),
            "genero": m.genero,
            "es_cabeza": es_cabeza,
        }
    
    def construir_familia_nuclear(padre_ids, grupo_ids, apellido_default=None):
        """
        Construye una familia nuclear desde un conjunto de padres.
        Retorna None si no hay hijos.
        """
        padres_lista = []
        hijos_lista = []
        
        # Construir lista de padres
        for pid in padre_ids:
            if pid not in miembros_map:
                continue
            m = miembros_map[pid]
            es_femenino = (m.genero or "").lower() in ("femenino", "f", "mujer")
            es_cabeza = es_cabeza_familia(m.id, grupo_ids)
            padres_lista.append(construir_datos_persona(
                m, 
                "Madre" if es_femenino else "Padre",
                es_cabeza
            ))
        
        # Buscar hijos directos
        hijos_ids = obtener_hijos_directos(padre_ids, grupo_ids)
        
        for hid in hijos_ids:
            if hid not in miembros_map:
                continue
            m = miembros_map[hid]
            es_femenino = (m.genero or "").lower() in ("femenino", "f", "mujer")
            hijos_lista.append(construir_datos_persona(
                m,
                "Hija" if es_femenino else "Hijo",
                False
            ))
        
        # Si no hay hijos, no es familia nuclear
        if not padres_lista or not hijos_lista:
            return None
        
        # Ordenar
        padres_lista.sort(key=lambda x: (0 if x["es_cabeza"] else 1, -(x["edad"] or 0)))
        hijos_lista.sort(key=lambda x: -(x["edad"] or 0))
        
        # Asegurar que hay cabeza
        todos = padres_lista + hijos_lista
        if not any(p["es_cabeza"] for p in todos):
            padres_lista[0]["es_cabeza"] = True
        
        # Determinar apellido
        if apellido_default:
            apellido = apellido_default
        else:
            # Usar el apellido del primer padre
            primer_padre_id = list(padre_ids)[0]
            if primer_padre_id in miembros_map:
                apellido = miembros_map[primer_padre_id].apellidos
            else:
                apellido = "Familia"
        
        # Determinar nombre de familia (por el padre/madre principal)
        nombre_familia = None
        for p in padres_lista:
            if p["es_cabeza"]:
                nombre_familia = p["nombre_completo"].split()[0]  # Primer nombre
                break
        if not nombre_familia and padres_lista:
            nombre_familia = padres_lista[0]["nombre_completo"].split()[0]
        
        return {
            "apellido": apellido,
            "nombre_referencia": nombre_familia,  # "Familia de Jhon"
            "padres": padres_lista,
            "hijos": hijos_lista,
            "miembros": padres_lista + hijos_lista,
            "num_hijos": len(hijos_lista),
            "padre_ids": padre_ids,  # Para identificar duplicados
        }
    
    def extraer_todas_subfamilias_nucleares(grupo_ids):
        """
        Extrae TODAS las subfamilias nucleares de un grupo.
        Cada pareja con hijos = 1 subfamilia.
        Cada padre/madre soltero con hijos = 1 subfamilia.
        """
        subfamilias = []
        parejas_procesadas = set()
        padres_solteros_procesados = set()
        
        # 1) Parejas con hijos
        parejas = obtener_parejas_en_grupo(grupo_ids)
        for p1, p2 in parejas:
            key = tuple(sorted([p1, p2]))
            if key in parejas_procesadas:
                continue
            
            sf = construir_familia_nuclear({p1, p2}, grupo_ids)
            if sf:
                subfamilias.append(sf)
                parejas_procesadas.add(key)
        
        # 2) Padres/madres solteros (no en pareja) con hijos
        for pid in grupo_ids:
            # Si ya es parte de una pareja, saltar
            es_parte_de_pareja = any(pid in {p1, p2} for p1, p2 in parejas)
            if es_parte_de_pareja:
                continue
            
            if pid in padres_solteros_procesados:
                continue
            
            # Ver si tiene hijos directos
            hijos = obtener_hijos_directos({pid}, grupo_ids)
            if hijos:
                sf = construir_familia_nuclear({pid}, grupo_ids)
                if sf:
                    subfamilias.append(sf)
                    padres_solteros_procesados.add(pid)
        
        return subfamilias
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CLASIFICAR GRUPOS
    # ═══════════════════════════════════════════════════════════════════════════
    familias_extendidas = []
    familias_nucleares = []  # ← TODAS las nucleares, incluso las de extendidas
    parejas = []
    
    # Para evitar duplicados en nucleares
    nucleares_agregadas = set()  # Guardamos frozenset de padre_ids
    
    total_hijos = 0
    
    for grupo_ids in grupos:
        miembros_grupo = [miembros_map[mid] for mid in grupo_ids if mid in miembros_map]
        
        if len(miembros_grupo) < 2:
            continue
        
        # Apellido principal del grupo
        apellidos_count = defaultdict(int)
        for m in miembros_grupo:
            apellidos_count[m.apellidos] += 1
        apellido_principal = max(apellidos_count, key=apellidos_count.get)
        
        # Teléfono y dirección
        telefono_familia = None
        direccion_familia = None
        for m in miembros_grupo:
            if m.telefono and not telefono_familia:
                telefono_familia = m.telefono
            if hasattr(m, 'direccion') and m.direccion and not direccion_familia:
                direccion_familia = m.direccion
        
        # Extraer TODAS las subfamilias nucleares del grupo
        subfamilias = extraer_todas_subfamilias_nucleares(grupo_ids)
        parejas_grupo = obtener_parejas_en_grupo(grupo_ids)
        
        # ═══════════════════════════════════════════════════════════════════════
        # AGREGAR TODAS LAS NUCLEARES A LA LISTA (siempre)
        # ═══════════════════════════════════════════════════════════════════════
        for sf in subfamilias:
            key = frozenset(sf["padre_ids"])
            if key not in nucleares_agregadas:
                sf["telefono"] = telefono_familia
                sf["direccion"] = direccion_familia
                familias_nucleares.append(sf)
                nucleares_agregadas.add(key)
                total_hijos += sf["num_hijos"]
        
        # ═══════════════════════════════════════════════════════════════════════
        # CLASIFICAR EL GRUPO COMPLETO
        # ═══════════════════════════════════════════════════════════════════════
        
        # CASO 1: Múltiples subfamilias = Familia Extendida
        if len(subfamilias) >= 2:
            familias_extendidas.append({
                "apellido": apellido_principal,
                "subfamilias": subfamilias,
                "total_miembros": len(miembros_grupo),
                "num_nucleos": len(subfamilias),
                "telefono": telefono_familia,
                "direccion": direccion_familia,
            })
            continue
        
        # CASO 2: Una sola subfamilia = Ya se agregó a nucleares, no va a extendidas
        if len(subfamilias) == 1:
            continue  # Ya está en familias_nucleares
        
        # CASO 3: Solo pareja (sin hijos)
        if len(miembros_grupo) == 2 and parejas_grupo:
            p1_id, p2_id = parejas_grupo[0]
            p1 = miembros_map.get(p1_id)
            p2 = miembros_map.get(p2_id)
            
            if p1 and p2:
                parejas.append({
                    "persona1": {
                        "id": p1.id,
                        "nombre_completo": f"{p1.nombres} {p1.apellidos}",
                        "edad": p1.calcular_edad(),
                        "genero": p1.genero,
                    },
                    "persona2": {
                        "id": p2.id,
                        "nombre_completo": f"{p2.nombres} {p2.apellidos}",
                        "edad": p2.calcular_edad(),
                        "genero": p2.genero,
                    },
                    "telefono": telefono_familia,
                    "anios_juntos": None,
                })
            continue
        
        # CASO 4: Grupo conectado sin estructura nuclear clara (hermanos, primos, etc.)
        # Lo mostramos como familia extendida general
        if len(miembros_grupo) >= 2:
            miembros_lista = []
            for m in miembros_grupo:
                rol = "Familiar"
                for otro_id in grupo_ids:
                    if otro_id == m.id:
                        continue
                    key = (otro_id, m.id)
                    if key in relaciones_info:
                        tipo = relaciones_info[key]["tipo"]
                        rol = obtener_rol_genero(tipo, m.genero)
                        break
                
                miembros_lista.append(construir_datos_persona(
                    m, rol, es_cabeza_familia(m.id, grupo_ids)
                ))
            
            miembros_lista.sort(key=lambda x: -(x["edad"] or 0))
            if not any(x["es_cabeza"] for x in miembros_lista):
                miembros_lista[0]["es_cabeza"] = True
            
            familias_extendidas.append({
                "apellido": apellido_principal,
                "subfamilias": [],
                "miembros_generales": miembros_lista,
                "total_miembros": len(miembros_grupo),
                "num_nucleos": 0,
                "telefono": telefono_familia,
                "direccion": direccion_familia,
                "es_grupo_general": True,
            })
    
    # ═══════════════════════════════════════════════════════════════════════════
    # APLICAR FILTRO POR TIPO
    # ═══════════════════════════════════════════════════════════════════════════
    if tipo_filtro == "extendida":
        familias_nucleares = []
        parejas = []
    elif tipo_filtro == "nuclear":
        familias_extendidas = []
        parejas = []
    elif tipo_filtro == "pareja":
        familias_extendidas = []
        familias_nucleares = []
    
    # ═══════════════════════════════════════════════════════════════════════════
    # ORDENAR
    # ═══════════════════════════════════════════════════════════════════════════
    familias_extendidas.sort(key=lambda f: f["apellido"].lower())
    familias_nucleares.sort(key=lambda f: f["apellido"].lower())
    parejas.sort(key=lambda p: p["persona1"]["nombre_completo"].lower())
    
    # ═══════════════════════════════════════════════════════════════════════════
    # ESTADÍSTICAS
    # ═══════════════════════════════════════════════════════════════════════════
    total_personas_extendidas = sum(f.get("total_miembros", 0) for f in familias_extendidas)
    total_personas_nucleares = sum(len(f["miembros"]) for f in familias_nucleares)
    total_personas_parejas = len(parejas) * 2
    
    total_resultados = len(familias_extendidas) + len(familias_nucleares) + len(parejas)
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CONTEXTO
    # ═══════════════════════════════════════════════════════════════════════════
    context = {
        "familias_extendidas": familias_extendidas,
        "familias_nucleares": familias_nucleares,
        "parejas": parejas,
        
        "total_familias_extendidas": len(familias_extendidas),
        "total_familias_nucleares": len(familias_nucleares),
        "total_parejas": len(parejas),
        "total_personas_en_familias": total_personas_extendidas + total_personas_nucleares + total_personas_parejas,
        "total_hijos": total_hijos,
        "total_resultados": total_resultados,
        
        "query": query,
        "tipo_filtro": tipo_filtro,
        
        "hoy": hoy,
        "CFG": CFG,
    }
    
    return render(
        request,
        "miembros_app/reportes/reporte_relaciones_familiares.html",
        context
    )

# -------------------------------------
# REPORTE: CUMPLEAÑOS DEL MES
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_cumple_mes(request):
    """
    Reporte imprimible de los cumpleaños de un mes.
    - Por defecto muestra el mes actual.
    - Filtros:
        * solo_activos: solo miembros activos en Torre Fuerte.
        * solo_oficiales: solo mayores de edad mínima oficial.
    """

    hoy = timezone.localdate()
    edad_minima = get_edad_minima_miembro_oficial()

    # --- Mes seleccionado ---
    mes_str = request.GET.get("mes", "").strip()
    if mes_str.isdigit():
        mes = int(mes_str)
        if mes < 1 or mes > 12:
            mes = hoy.month
    else:
        mes = hoy.month

    # Año solo para mostrar en el título (no afecta el filtro)
    anio = hoy.year

    # Flags de filtros (por defecto: solo_activos=ON, solo_oficiales=OFF)
    solo_activos = request.GET.get("solo_activos", "1") == "1"
    solo_oficiales = request.GET.get("solo_oficiales", "0") == "1"

    # Nombres de meses
    MESES_ES = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    nombre_mes = MESES_ES.get(mes, "")

    # Base: miembros con fecha de nacimiento en ese mes
    miembros = Miembro.objects.filter(
        fecha_nacimiento__isnull=False,
        fecha_nacimiento__month=mes,
    )

    # Solo activos (siguen en la iglesia)
    if solo_activos:
        miembros = miembros.filter(activo=True)

    # Solo oficiales (>= edad mínima)
    if solo_oficiales:
        cutoff = hoy - timedelta(days=edad_minima * 365)
        miembros = miembros.filter(fecha_nacimiento__lte=cutoff)

    # Añadimos el día del mes y ordenamos
    miembros = (
        miembros
        .annotate(dia=ExtractDay("fecha_nacimiento"))
        .order_by("dia", "apellidos", "nombres")
    )

    # Calculamos cuántos años cumplen (edad_que_cumple)
    for m in miembros:
        if m.fecha_nacimiento:
            edad_actual = m.calcular_edad()
            if edad_actual is not None:
                m.edad_que_cumple = edad_actual + 1
            else:
                m.edad_que_cumple = None
        else:
            m.edad_que_cumple = None

    context = {
        "miembros": miembros,
        "mes": mes,
        "anio": anio,
        "nombre_mes": nombre_mes,
        "solo_activos": solo_activos,
        "solo_oficiales": solo_oficiales,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/reportes/cumple_mes.html", context)


# -------------------------------------
# REPORTE: MIEMBROS NUEVOS DEL MES
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_miembros_nuevos_mes(request):
    """
    Reporte de miembros que ingresaron a la iglesia en un mes concreto.
    - Por defecto muestra el mes actual.
    - Filtros:
        * solo_activos: solo miembros activos en Torre Fuerte.
        * solo_oficiales: solo mayores de edad mínima oficial.
    """

    hoy = timezone.localdate()
    edad_minima = get_edad_minima_miembro_oficial()

    # --- Mes seleccionado (input type="month" -> YYYY-MM) ---
    mes_str = request.GET.get("mes", "").strip()

    if mes_str:
        # Intentamos parsear "YYYY-MM"
        try:
            partes = mes_str.split("-")
            anio = int(partes[0])
            mes = int(partes[1])
            if mes < 1 or mes > 12:
                raise ValueError
        except Exception:
            anio = hoy.year
            mes = hoy.month
            mes_str = f"{anio:04d}-{mes:02d}"
    else:
        anio = hoy.year
        mes = hoy.month
        mes_str = f"{anio:04d}-{mes:02d}"

    # --- Flags de filtros (por defecto: solo_activos=ON, solo_oficiales=OFF) ---
    solo_activos = request.GET.get("solo_activos", "1") == "1"
    solo_oficiales = request.GET.get("solo_oficiales", "0") == "1"

    # Texto de búsqueda
    query = request.GET.get("q", "").strip()

    # Nombres de meses (para mostrar título bonito)
    MESES_ES = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    nombre_mes = MESES_ES.get(mes, "")

    # --- Rango de fechas del mes seleccionado ---
    # Primer día del mes
    fecha_inicio = date(anio, mes, 1)
    # Primer día del mes siguiente
    if mes == 12:
        fecha_fin_mes_siguiente = date(anio + 1, 1, 1)
    else:
        fecha_fin_mes_siguiente = date(anio, mes + 1, 1)
    # Último día del mes seleccionado
    fecha_fin = fecha_fin_mes_siguiente - timedelta(days=1)

    # Base: miembros con fecha de ingreso en ese rango
    miembros = Miembro.objects.filter(
        fecha_ingreso_iglesia__isnull=False,
        fecha_ingreso_iglesia__gte=fecha_inicio,
        fecha_ingreso_iglesia__lte=fecha_fin,
    )

    # Solo activos (siguen en la iglesia)
    if solo_activos:
        miembros = miembros.filter(activo=True)

    # Solo oficiales (>= edad mínima)
    if solo_oficiales:
        cutoff = hoy - timedelta(days=edad_minima * 365)
        miembros = miembros.filter(
            Q(fecha_nacimiento__isnull=False) & Q(fecha_nacimiento__lte=cutoff)
        )

    # Búsqueda por nombre, apellidos, correo o teléfono
    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query)
            | Q(apellidos__icontains=query)
            | Q(email__icontains=query)
            | Q(telefono__icontains=query)
            | Q(telefono_secundario__icontains=query)
        )

    # Orden: por fecha de ingreso y luego por nombre
    miembros = miembros.order_by(
        "fecha_ingreso_iglesia",
        "apellidos",
        "nombres",
    )

    # Calculamos la edad actual (para mostrar en la tabla)
    for m in miembros:
        if hasattr(m, "calcular_edad"):
            m.edad_actual = m.calcular_edad()
        else:
            if m.fecha_nacimiento:
                fn = m.fecha_nacimiento
                edad = hoy.year - fn.year
                if (hoy.month, hoy.day) < (fn.month, fn.day):
                    edad -= 1
                m.edad_actual = edad
            else:
                m.edad_actual = None

    context = {
        "miembros": miembros,
        "mes_str": mes_str,
        "anio": anio,
        "mes": mes,
        "nombre_mes": nombre_mes,
        "solo_activos": solo_activos,
        "solo_oficiales": solo_oficiales,
        "query": query,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(
        request,
        "miembros_app/reportes/reporte_miembros_nuevos_mes.html",
        context,
    )


# -------------------------------
# CARTA DE SALIDA / TRASLADO
# -------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def carta_salida_miembro(request, pk):
    """
    Genera una carta imprimible de salida / traslado para un miembro,
    PERO solo si la razón de salida lo permite.
    """
    try:
        miembro = get_object_or_404(Miembro, pk=pk)

        # ✅ BLOQUEO: solo permitir carta si la razón lo permite
        if (not miembro.razon_salida) or (not getattr(miembro.razon_salida, "permite_carta", False)):
            messages.error(
                request,
                "No aplica carta para este Miembro. "
                
            )
            return redirect("miembros_app:inactivo_detalle", pk=miembro.pk)

        hoy = timezone.localdate()

        iglesia_nombre = getattr(settings, "NOMBRE_IGLESIA", "Iglesia Torre Fuerte")
        iglesia_ciudad = getattr(settings, "CIUDAD_IGLESIA", "Higüey, República Dominicana")
        pastor_principal = getattr(settings, "PASTOR_PRINCIPAL", "Pastor de la iglesia")

        context = {
            "miembro": miembro,
            "hoy": hoy,
            "iglesia_nombre": iglesia_nombre,
            "iglesia_ciudad": iglesia_ciudad,
            "pastor_principal": pastor_principal,
        }

        return render(
            request,
            "miembros_app/cartas/carta_salida.html",
            context,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f"<h2>Error en carta_salida_miembro</h2>"
            f"<pre>{e}</pre>",
            status=500,
        )

from django.utils import timezone  # ya lo tienes arriba

@login_required
@permission_required("miembros_app.add_miembro", raise_exception=True)
def nuevo_creyente_crear(request):
    """
    Registro rápido de nuevos creyentes.
    Guarda en Miembro pero marcando nuevo_creyente=True
    y sin mezclarlos todavía con la membresía oficial.
    """

    def wa_digits(v):
        # deja solo números (WhatsApp wa.me no acepta + ni espacios)
        return "".join(ch for ch in (v or "") if ch.isdigit())

    if request.method == "POST":
        form = NuevoCreyenteForm(request.POST)
        if form.is_valid():
            miembro = form.save()

            # 🔔 Crear notificación de nuevo creyente
            try:
                url_detalle = reverse(
                    "miembros_app:nuevo_creyente_editar",
                    args=[miembro.pk]
                )

                crear_notificacion(
                    usuario=request.user,
                    titulo="Nuevo creyente registrado",
                    mensaje=f"{miembro.nombres} {miembro.apellidos} ha entregado su vida a Cristo.",
                    url_name=url_detalle,
                    tipo="info",
                )
            except Exception as e:
                print("Error creando notificación de nuevo creyente:", e)

            # ✅ Mensaje de éxito
            messages.success(
                request,
                f"Nuevo creyente registrado correctamente: "
                f"{miembro.nombres} {miembro.apellidos}.",
            )

            # ───────────────────────────────────────────
            # 📲 WhatsApp (opcional, desde el formulario)
            # ───────────────────────────────────────────
            enviar_whatsapp = request.POST.get("enviar_whatsapp") == "1"

            if enviar_whatsapp and miembro.whatsapp:
                mensaje = (
                    f"Hola {miembro.nombres}, 👋\n\n"
                    "¡Bienvenido a Iglesia Torre Fuerte! 🙏\n\n"
                    "Nos alegra mucho que hayas dado este paso tan importante. "
                    "Muy pronto alguien del equipo se pondrá en contacto contigo "
                    "para acompañarte en este nuevo comienzo.\n\n"
                    "Dios te bendiga."
                )

                wa_url = (
                    f"https://wa.me/{wa_digits(miembro.whatsapp)}"
                    f"?text={quote(mensaje)}"
                )

                # 👉 Redirige directamente a WhatsApp
                return redirect(wa_url)

            # 🔁 Flujo normal (si no se envía WhatsApp)
            return redirect("miembros_app:nuevo_creyente_lista")

    else:
        form = NuevoCreyenteForm()

    context = {
        "form": form,
        "modo": "crear",
    }
    return render(request, "miembros_app/nuevos_creyentes_form.html", context)


@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def nuevo_creyente_lista(request):

    """
    Lista de nuevos creyentes (miembros con nuevo_creyente=True),
    con filtros por texto, género, rango de fecha de conversión
    y si tienen o no datos de contacto.
    """

    query = request.GET.get("q", "").strip()
    genero_filtro = request.GET.get("genero", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    solo_contacto = request.GET.get("solo_contacto", "") == "1"

    ver_inactivos = request.GET.get("ver_inactivos", "") == "1"

    miembros = Miembro.objects.filter(nuevo_creyente=True)

    # ✅ Por defecto: SOLO activos
    if not ver_inactivos:
        miembros = miembros.filter(activo=True)


    # Búsqueda por nombre / contacto
    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query)
            | Q(apellidos__icontains=query)
            | Q(telefono__icontains=query)
            | Q(telefono_secundario__icontains=query)
            | Q(email__icontains=query)
        )

    # Filtro por género
    if genero_filtro:
        miembros = miembros.filter(genero=genero_filtro)

    # Filtro por rango de fecha de conversión
    if fecha_desde:
        try:
            miembros = miembros.filter(fecha_conversion__gte=fecha_desde)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            miembros = miembros.filter(fecha_conversion__lte=fecha_hasta)
        except ValueError:
            pass

    # Solo los que tienen algún dato de contacto
    if solo_contacto:
        miembros = miembros.filter(
            Q(telefono__isnull=False, telefono__gt="")
            | Q(telefono_secundario__isnull=False, telefono_secundario__gt="")
            | Q(email__isnull=False, email__gt="")
        )

    # ✅ Marcar si ya fue enviado al módulo Nuevo Creyente (tiene expediente)
    miembros = miembros.annotate(
        nc_enviado=Exists(
            NuevoCreyenteExpediente.objects.filter(miembro_id=OuterRef("pk"))
        )
    )
        

    # ✅ Marcar si tiene Padres Espirituales asignados (ManyToMany)
    miembros = miembros.annotate(
        padres_espirituales_count=Count("padres_espirituales", distinct=True)
    )

    miembros = miembros.order_by(
        "-fecha_conversion",
        "-fecha_creacion",
        "apellidos",
        "nombres",
        
    )

    # Para llenar el select de género en la plantilla
    generos_choices = Miembro._meta.get_field("genero").choices

    context = {
        "miembros": miembros,
        "query": query,
        "genero_filtro": genero_filtro,
        "generos_choices": generos_choices,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "solo_contacto": solo_contacto,
        "hoy": timezone.localdate(),
        "ver_inactivos": ver_inactivos,

    }
    return render(
        request,
        "miembros_app/nuevos_creyentes_lista.html",
        context,
    )
@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
def nuevo_creyente_editar(request, pk):

    """
    Editar un nuevo creyente usando el mismo formulario sencillo.
    Solo permite editar registros marcados como nuevo_creyente=True.
    """
    miembro = get_object_or_404(Miembro, pk=pk, nuevo_creyente=True)

    if request.method == "POST":
        form = NuevoCreyenteForm(request.POST, instance=miembro)
        if form.is_valid():
            miembro = form.save()
            messages.success(
                request,
                f"Datos del nuevo creyente actualizados: {miembro.nombres} {miembro.apellidos}."
            )
            return redirect("miembros_app:nuevo_creyente_lista")
    else:
        form = NuevoCreyenteForm(instance=miembro)

    context = {
        "form": form,
        "modo": "editar",  # 👈 así el template sabe que está en modo edición
        "miembro": miembro,
    }
    return render(request, "miembros_app/nuevos_creyentes_form.html", context)
# -------------------------------------
# REPORTE: NUEVOS CREYENTES
# -------------------------------------
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_nuevos_creyentes(request):

    """
    Reporte imprimible de nuevos creyentes.
    Permite filtrar por nombre/contacto y por rango de fecha de conversión.
    """
    query = request.GET.get("q", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    solo_contacto = request.GET.get("solo_contacto", "") == "1"

    miembros = Miembro.objects.filter(nuevo_creyente=True)

    # --- Búsqueda general ---
    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query)
            | Q(apellidos__icontains=query)
            | Q(telefono__icontains=query)
            | Q(telefono_secundario__icontains=query)
            | Q(email__icontains=query)
        )

    # --- Filtro por fechas ---
    if fecha_desde:
        try:
            miembros = miembros.filter(fecha_conversion__gte=fecha_desde)
        except:
            pass

    if fecha_hasta:
        try:
            miembros = miembros.filter(fecha_conversion__lte=fecha_hasta)
        except:
            pass

    # --- Solo los que tienen algún contacto ---
    if solo_contacto:
        miembros = miembros.filter(
            Q(telefono__isnull=False, telefono__gt="")
            | Q(telefono_secundario__isnull=False, telefono_secundario__gt="")
            | Q(email__isnull=False, email__gt="")
        )

    miembros = miembros.order_by(
        "-fecha_conversion",
        "-fecha_creacion",
        "apellidos",
        "nombres",
    )

    context = {
        "miembros": miembros,
        "query": query,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "solo_contacto": solo_contacto,
        "hoy": timezone.localdate(),
    }

    return render(
        request,
        "miembros_app/reportes/reporte_nuevos_creyentes.html",
        context,
    )
@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def nuevo_creyente_ficha(request, pk):
    """
    Ficha imprimible del nuevo creyente.
    """
    miembro = get_object_or_404(Miembro, pk=pk, nuevo_creyente=True)

    # 🔍 Verificar si tiene expediente ABIERTO
    expediente_abierto = NuevoCreyenteExpediente.objects.filter(
        miembro=miembro,
        estado="abierto"
    ).exists()

    context = {
        "miembro": miembro,
        "hoy": timezone.localdate(),
        # 👉 solo se puede dar salida si NO hay expediente abierto
        "puede_dar_salida": not expediente_abierto,
        "expediente_abierto": expediente_abierto,
    }

    return render(
        request,
        "miembros_app/reportes/nuevo_creyente_ficha.html",
        context,
    )
def generar_pdf_ficha_miembro(miembro):
    """
    Genera un PDF de la ficha pastoral del miembro y devuelve
    la ruta absoluta del archivo generado.
    """
    from django.conf import settings

    # Configuración general (para CFG en la plantilla)
    config = ConfiguracionSistema.load()

    # Traer familiares igual que en la ficha imprimible
    relaciones_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro=miembro) | Q(familiar=miembro))
        .select_related("miembro", "familiar")
    )

    relaciones_familia = []
    parejas_vistas = set()

    for rel in relaciones_qs:
        if rel.tipo_relacion == "conyuge":
            pareja = frozenset({rel.miembro_id, rel.familiar_id})
            if pareja in parejas_vistas:
                continue
            parejas_vistas.add(pareja)
            relaciones_familia.append(rel)
        else:
            relaciones_familia.append(rel)

    edad_minima = get_edad_minima_miembro_oficial()

    # ✨ AQUÍ SÍ INCLUIMOS CFG EN EL CONTEXTO
    context = {
        "miembro": miembro,
        "relaciones_familia": relaciones_familia,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
        "CFG": config,  # ← ESTA ES LA CLAVE QUE FALTABA
    }

    # Renderizamos la misma plantilla que usas para imprimir la ficha
    html = render_to_string("miembros_app/reportes/miembro_ficha.html", context)

    # Creamos el PDF en memoria
    resultado = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=resultado)

    if pisa_status.err:
        raise Exception("Error generando el PDF de la ficha del miembro.")

    # Carpeta donde guardaremos las fichas
    carpeta = os.path.join(settings.MEDIA_ROOT, "fichas_miembro")
    os.makedirs(carpeta, exist_ok=True)

    nombre_archivo = f"ficha_miembro_{miembro.pk}.pdf"
    ruta_pdf = os.path.join(carpeta, nombre_archivo)

    # Guardamos el contenido del PDF en disco
    with open(ruta_pdf, "wb") as f:
        f.write(resultado.getvalue())

    return ruta_pdf

@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def listado_miembros_enviar_email(request):

    """
    Genera un PDF EXACTO del listado de miembros usando la URL real
    y lo envía por correo.
    """
    config = ConfiguracionSistema.load()

    # Obtener los filtros actuales
    filtros = request.GET.urlencode()
    base_url = request.build_absolute_uri(reverse("miembros_app:reporte_listado_miembros"))

    # Construir URL completa con filtros
    url_completa = base_url + (f"?{filtros}" if filtros else "")

    # Datos iniciales del formulario
    email_default = config.email_oficial or settings.DEFAULT_FROM_EMAIL
    asunto_default = "Listado general de miembros"
    mensaje_default = f"Adjunto el listado general de miembros de {config.nombre_iglesia or 'nuestra iglesia'}."

    if request.method == "POST":
        form = EnviarFichaMiembroEmailForm(request.POST)
        if form.is_valid():
            destinatario = form.cleaned_data["destinatario"]
            asunto = form.cleaned_data["asunto"] or asunto_default
            mensaje = form.cleaned_data["mensaje"] or mensaje_default

            try:
                # GENERAR PDF USANDO LA URL REAL
                pdf_bytes = generar_pdf_desde_url(url_completa)

                # Cuerpo del correo
                body_html = (
                    f"<p>{mensaje}</p>"
                    "<p style='margin-top:16px;'>Bendiciones,<br><strong>Soid_Tf_2</strong></p>"
                )

                # Enviar correo con adjunto
                enviar_correo_sistema(
                    subject=asunto,
                    heading="Listado general de miembros",
                    body_html=body_html,
                    destinatarios=destinatario,
                    meta_text="Correo enviado desde Soid_Tf_2",
                    extra_context={"CFG": config},
                    adjuntos=[("listado_miembros.pdf", pdf_bytes)],
                )

                messages.success(
                    request, f"Correo enviado correctamente a {destinatario}"
                )
                return redirect("miembros_app:reporte_listado_miembros")

            except Exception as e:
                messages.error(request, f"No se pudo enviar el correo: {e}")
                return redirect("miembros_app:reporte_listado_miembros")

    else:
        form = EnviarFichaMiembroEmailForm(
            initial={
                "destinatario": email_default,
                "asunto": asunto_default,
                "mensaje": mensaje_default,
            }
        )

    return render(
        request,
        "core/enviar_email.html",
        {
            "form": form,
            "titulo_pagina": "Enviar listado por correo",
            "descripcion": "Se generará un PDF idéntico al de impresión usando Chrome Headless.",
            "objeto_label": "Listado general de miembros",
            "url_cancelar": reverse("miembros_app:reporte_listado_miembros"),
            "adjunto_auto_nombre": "listado_miembros.pdf",
        },
    )




def generar_pdf_listado_miembros(request, miembros, filtros_context):
    """
    Genera un PDF usando la MISMA plantilla del listado,
    conservando el diseño original sin modificar nada.
    """
    config = ConfiguracionSistema.load()
    edad_minima = get_edad_minima_miembro_oficial()

    # Contexto igualito al de la vista original
    context = {
        "miembros": miembros,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
        "CFG": config,
        "hoy": timezone.localdate(),
    }
    context.update(filtros_context)

    # Renderizamos el MISMO HTML del listado
    html_string = render_to_string(
        "miembros_app/reportes/listado_miembros.html",
        context
    )

    # Convertir el HTML a PDF
    pdf_file = BytesIO()
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(pdf_file)

    # Guardar en MEDIA/reportes_miembros/
    carpeta = os.path.join(settings.MEDIA_ROOT, "reportes_miembros")
    os.makedirs(carpeta, exist_ok=True)

    ruta_pdf = os.path.join(carpeta, "listado_miembros.pdf")

    with open(ruta_pdf, "wb") as f:
        f.write(pdf_file.getvalue())

    return ruta_pdf
def generar_pdf_chrome_headless(request, html_content):
    """
    Genera un PDF en memoria usando Chrome Headless.
    El resultado es idéntico al PDF que genera el navegador.
    """
    chrome_path = get_chrome_path()

    # Crear archivo HTML temporal
    with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as html_temp:
        html_temp.write(html_content.encode("utf-8"))
        html_temp_path = html_temp.name

    # Archivo PDF temporal (salida)
    pdf_temp_path = html_temp_path.replace(".html", ".pdf")

    comando = [
        chrome_path,
        "--headless",
        "--disable-gpu",
        f"--print-to-pdf={pdf_temp_path}",
        html_temp_path,
    ]

    resultado = subprocess.run(
        comando,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    if resultado.returncode != 0:
        raise RuntimeError(
            f"Error ejecutando Chrome Headless (código {resultado.returncode}): {resultado.stderr}"
        )

    with open(pdf_temp_path, "rb") as f:
        pdf_bytes = f.read()

    os.remove(html_temp_path)
    os.remove(pdf_temp_path)

    return pdf_bytes
def generar_pdf_desde_url(url):
    """
    Genera un PDF desde una URL real usando Chrome Headless.
    El PDF es idéntico al generado desde 'Imprimir → Guardar como PDF'.
    """
    chrome_path = get_chrome_path()

    # Crear archivo PDF temporal
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as pdf_temp:
        pdf_temp_path = pdf_temp.name

    comando = [
        chrome_path,
        "--headless",
        "--disable-gpu",
        f"--print-to-pdf={pdf_temp_path}",
        url,  # aquí va la URL, no html_temp_path
    ]

    resultado = subprocess.run(
        comando,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    if resultado.returncode != 0:
        raise RuntimeError(
            f"Error ejecutando Chrome Headless (código {resultado.returncode}): {resultado.stderr}"
        )

    with open(pdf_temp_path, "rb") as f:
        pdf_bytes = f.read()

    os.remove(pdf_temp_path)

    return pdf_bytes


def get_chrome_path():
    """
    Devuelve una ruta válida al ejecutable de Chrome/Chromium.
    Orden de prioridad:
    1) settings.CHROME_PATH
    2) variable de entorno CHROME_PATH
    3) detección automática (rutas típicas y ejecutables conocidos)
    """
    global CHROME_PATH

    # 1) Si ya está resuelto y existe, úsalo
    if CHROME_PATH and os.path.exists(CHROME_PATH):
        return CHROME_PATH

    # 2) Revisar settings.CHROME_PATH
    settings_path = getattr(settings, "CHROME_PATH", "") or ""
    if settings_path and os.path.exists(settings_path):
        CHROME_PATH = settings_path
        return CHROME_PATH

    # 3) Revisar variable de entorno CHROME_PATH
    env_path = os.environ.get("CHROME_PATH")
    if env_path and os.path.exists(env_path):
        CHROME_PATH = env_path
        return CHROME_PATH

    # 4) Detección automática según sistema operativo
    system = platform.system()

    candidatos = []
    if system == "Windows":
        candidatos = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            shutil.which("chrome"),
            shutil.which("msedge"),  # Edge (Chromium) también sirve
        ]
    else:
        candidatos = [
            shutil.which("google-chrome"),
            shutil.which("google-chrome-stable"),
            shutil.which("chromium"),
            shutil.which("chromium-browser"),
        ]

    for ruta in candidatos:
        if ruta and os.path.exists(ruta):
            CHROME_PATH = ruta
            return CHROME_PATH

    # 5) Si no se encuentra nada, error real
    raise RuntimeError(
        "No se encontró Chrome/Chromium. "
        "Instálalo o define CHROME_PATH en settings.py o en variables de entorno."
    )
@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def nuevos_creyentes_enviar_email(request):

    """
    Genera un PDF del listado de nuevos creyentes (con los filtros actuales)
    y lo envía por correo.
    """
    config = ConfiguracionSistema.load()

    # Tomamos los filtros que estén activos en la URL
    filtros = request.GET.urlencode()

    # OJO: el name correcto de la ruta es 'nuevo_creyente_lista'
    base_url = request.build_absolute_uri(
        reverse("miembros_app:nuevo_creyente_lista")
    )
    url_completa = base_url + (f"?{filtros}" if filtros else "")

    email_default = config.email_oficial or settings.DEFAULT_FROM_EMAIL
    asunto_default = "Listado de nuevos creyentes"
    mensaje_default = (
        f"Adjunto el listado de nuevos creyentes de "
        f"{config.nombre_iglesia or 'nuestra iglesia'}."
    )

    if request.method == "POST":
        form = EnviarFichaMiembroEmailForm(request.POST)
        if form.is_valid():
            destinatario = form.cleaned_data["destinatario"]
            asunto = form.cleaned_data["asunto"] or asunto_default
            mensaje = form.cleaned_data["mensaje"] or mensaje_default

            try:
                # Generar el PDF desde la URL real (Chrome Headless)
                pdf_bytes = generar_pdf_desde_url(url_completa)

                body_html = (
                    f"<p>{mensaje}</p>"
                    "<p style='margin-top:16px;'>"
                    "Bendiciones,<br><strong>Soid_Tf_2</strong>"
                    "</p>"
                )

                enviar_correo_sistema(
                    subject=asunto,
                    heading="Listado de nuevos creyentes",
                    body_html=body_html,
                    destinatarios=destinatario,
                    meta_text="Correo enviado desde Soid_Tf_2",
                    extra_context={"CFG": config},
                    adjuntos=[("nuevos_creyentes.pdf", pdf_bytes)],
                )

                messages.success(
                    request,
                    f"Correo enviado correctamente a {destinatario}."
                )
                return redirect("miembros_app:nuevo_creyente_lista")

            except Exception as e:
                messages.error(
                    request,
                    f"No se pudo enviar el correo: {e}",
                )
                return redirect("miembros_app:nuevo_creyente_lista")
    else:
        form = EnviarFichaMiembroEmailForm(
            initial={
                "destinatario": email_default,
                "asunto": asunto_default,
                "mensaje": mensaje_default,
            }
        )

    return render(
        request,
        "core/enviar_email.html",
        {
            "form": form,
            "titulo_pagina": "Enviar listado de nuevos creyentes",
            "descripcion": "Se generará un PDF idéntico al de impresión usando Chrome Headless.",
            "objeto_label": "Listado de nuevos creyentes",
            "url_cancelar": reverse("miembros_app:nuevo_creyente_lista"),
            "adjunto_auto_nombre": "nuevos_creyentes.pdf",
        },
    )
@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def exportar_miembros_excel(request):

    """
    Exporta a Excel los miembros que se están viendo en el listado,
    respetando TODOS los filtros aplicados en la vista miembro_lista.
    """

    # Base: solo miembros que NO son nuevos creyentes
    miembros_base = Miembro.objects.filter(nuevo_creyente=False)

    # Reutilizamos exactamente la misma lógica de filtros
    miembros, filtros_context = filtrar_miembros(request, miembros_base)

    # Creamos el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Miembros"

    # Encabezados de columnas
    headers = [
        "ID",
        "Nombres",
        "Apellidos",
        "Edad",
        "Género",
        "Estado",
        "Categoría edad",
        "Teléfono",
        "Email",
        "Bautizado",
        "Fecha ingreso",
        "Activo",
    ]
    ws.append(headers)

    # Rellenar filas con los datos filtrados
    for m in miembros:
        # Edad calculada (si no hay fecha de nacimiento, queda vacío)
        try:
            edad = m.edad
        except Exception:
            edad = None

        # Género, estado y categoría (usamos los display si existen)
        genero_display = (
            m.get_genero_display() if hasattr(m, "get_genero_display") else m.genero
        )
        estado_display = (
            m.get_estado_miembro_display()
            if hasattr(m, "get_estado_miembro_display")
            else m.estado_miembro
        )
        categoria_display = (
            m.get_categoria_edad_display()
            if hasattr(m, "get_categoria_edad_display")
            else m.categoria_edad
        )

        # Fecha de ingreso formateada
        if getattr(m, "fecha_ingreso_iglesia", None):
            fecha_ingreso_str = m.fecha_ingreso_iglesia.strftime("%d/%m/%Y")
        else:
            fecha_ingreso_str = ""

        row = [
            m.id,
            m.nombres,
            m.apellidos,
            edad if edad is not None else "",
            genero_display or "",
            estado_display or "",
            categoria_display or "",
            m.telefono or "",
            m.email or "",
            "Sí" if m.bautizado_confirmado else "No",
            fecha_ingreso_str,
            "Activo" if m.activo else "Inactivo",
        ]
        ws.append(row)

    # Ajustar un poco el ancho de las columnas automáticamente
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # A, B, C...
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        # Limitar el ancho a algo razonable
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width

    # Preparar la respuesta HTTP con el archivo Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "miembros_filtrados.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
@require_POST
@permission_required("miembros_app.add_miembro", raise_exception=True)
def importar_miembros_excel(request):

    """
    Importa miembros desde un archivo de Excel.
    Crea NUEVOS miembros usando las columnas del archivo.
    
    Formato esperado (fila 1 = encabezados):
    - Nombres
    - Apellidos
    - Genero       (Masculino / Femenino)
    - Estado       (Activo, Pasivo, En observación, En disciplina, Descarriado, Catecúmeno)
    - Telefono
    - Email
    - Fecha_nacimiento (dd/mm/aaaa o yyyy-mm-dd)
    """

    if request.method != "POST":
        messages.error(request, "Método no permitido para importar.")
        return redirect("miembros_app:lista")

    archivo = request.FILES.get("archivo_excel")

    if not archivo:
        messages.error(request, "No se ha enviado ningún archivo de Excel.")
        return redirect("miembros_app:lista")

    try:
        wb = load_workbook(filename=archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"El archivo no parece ser un Excel válido: {e}")
        return redirect("miembros_app:lista")

    # Leemos la fila de encabezados (fila 1)
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        messages.error(request, "El archivo está vacío.")
        return redirect("miembros_app:lista")

    # Creamos un mapa nombre_columna -> índice
    header_map = {}
    for idx, name in enumerate(header_row):
        if name:
            key = str(name).strip().lower()
            header_map[key] = idx

    # Columnas mínimas obligatorias
    columnas_obligatorias = ["nombres", "apellidos"]
    faltantes = [col for col in columnas_obligatorias if col not in header_map]

    if faltantes:
        msg = (
            "Faltan columnas obligatorias en el encabezado: "
            + ", ".join(faltantes)
            + ". Asegúrate de tener al menos: Nombres y Apellidos."
        )
        messages.error(request, msg)
        return redirect("miembros_app:lista")

    creados = 0
    omitidos = 0

    # Mapeos para género y estado
    mapa_genero = {
        "masculino": "masculino",
        "m": "masculino",
        "hombre": "masculino",
        "male": "masculino",
        "femenino": "femenino",
        "f": "femenino",
        "mujer": "femenino",
        "female": "femenino",
    }

    mapa_estado = {
        "activo": "activo",
        "pasivo": "pasivo",
        "en observación": "observacion",
        "en observacion": "observacion",
        "observacion": "observacion",
        "observación": "observacion",
        "disciplina": "disciplina",
        "en disciplina": "disciplina",
        "descarriado": "descarriado",
        "catecumeno": "catecumeno",
        "catecúmeno": "catecumeno",
    }

    # Recorremos las filas de datos (desde la fila 2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Si la fila está completamente vacía, la saltamos
        if not row or all((cell is None or str(cell).strip() == "") for cell in row):
            continue

        def get_value(nombre_col):
            idx = header_map.get(nombre_col)
            if idx is None:
                return ""
            value = row[idx]
            return "" if value is None else str(value).strip()

        nombres = get_value("nombres")
        apellidos = get_value("apellidos")

        if not nombres and not apellidos:
            # No tiene ni nombres ni apellidos, omitimos
            omitidos += 1
            continue

        genero_val = ""
        if "genero" in header_map:
            genero_raw = get_value("genero").lower()
            genero_val = mapa_genero.get(genero_raw, "")

        estado_val = "activo"
        if "estado" in header_map:
            estado_raw = get_value("estado").lower()
            estado_val = mapa_estado.get(estado_raw, "activo")

        telefono = get_value("telefono")
        email = get_value("email")

        fecha_nacimiento = None
        if "fecha_nacimiento" in header_map:
            valor_fecha = row[header_map["fecha_nacimiento"]]
            if valor_fecha:
                # Puede venir como fecha de Excel o como texto
                if hasattr(valor_fecha, "year"):
                    # Ya es un objeto date/datetime
                    try:
                        # Si es datetime, extraemos solo la fecha
                        fecha_nacimiento = valor_fecha.date()
                    except AttributeError:
                        fecha_nacimiento = valor_fecha
                else:
                    # Intentamos parsear texto
                    texto_fecha = str(valor_fecha).strip()
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                        try:
                            fecha_nacimiento = datetime.strptime(texto_fecha, fmt).date()
                            break
                        except ValueError:
                            continue

        # Creamos el miembro
        try:
            Miembro.objects.create(
                nombres=nombres,
                apellidos=apellidos,
                genero=genero_val,
                telefono=telefono,
                email=email,
                fecha_nacimiento=fecha_nacimiento,
                estado_miembro=estado_val,
                # El resto de campos quedan por defecto
            )
            creados += 1
        except Exception:
            omitidos += 1
            continue

    messages.success(
        request,
        f"Importación completada. Miembros creados: {creados}. Filas omitidas: {omitidos}.",
    )
    return redirect("miembros_app:lista")




def _miembro_tiene_asignacion_en_unidades(miembro_obj):
    """
    True si el miembro está asignado a alguna unidad (UnidadCargo o UnidadMembresia).
    Intenta filtrar solo asignaciones vigentes si el modelo tiene campo de fin/activo.
    """
    if not _modulo_estructura_activo():
        return False

    modelos = ("UnidadMembresia", "UnidadCargo")

    for model_name in modelos:
        Modelo = _safe_get_model("estructura_app", model_name)
        if not Modelo:
            continue

        # Detectar automáticamente el FK a Miembro dentro del modelo
        fk_name = None
        for f in Modelo._meta.fields:
            if getattr(f, "remote_field", None) and f.remote_field and f.remote_field.model == Miembro:
                fk_name = f.name
                break

        if not fk_name:
            continue

        qs = Modelo.objects.filter(**{fk_name: miembro_obj})

        # Si hay un campo de fin, intentamos tomar solo "vigentes"
        for end_field in ("fecha_fin", "fecha_final", "fecha_hasta", "fecha_salida"):
            if hasattr(Modelo, end_field):
                qs = qs.filter(**{f"{end_field}__isnull": True})
                break

        # Si existe un boolean "activo" en el modelo, también lo usamos
        if hasattr(Modelo, "activo"):
            qs = qs.filter(activo=True)

        if qs.exists():
            return True

    return False


CEDULA_RE = re.compile(r"^\d{3}-\d{7}-\d$")

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def ajax_validar_cedula(request):

    """
    GET /miembros/ajax/validar-cedula/?cedula=000-0000000-0&pk=123
    Devuelve si la cédula ya existe (excluyendo el pk si viene).
    """
    cedula = (request.GET.get("cedula") or "").strip()
    pk = (request.GET.get("pk") or "").strip()

    if not cedula:
        return JsonResponse({"ok": True, "empty": True, "valid_format": False, "exists": False})

    valid_format = bool(CEDULA_RE.match(cedula))
    if not valid_format:
        return JsonResponse({
            "ok": True,
            "empty": False,
            "valid_format": False,
            "exists": False,
            "message": "Formato inválido. Usa 000-0000000-0",
        })

    qs = Miembro.objects.filter(cedula=cedula)
    if pk.isdigit():
        qs = qs.exclude(pk=int(pk))

    exists = qs.exists()

    return JsonResponse({
        "ok": True,
        "empty": False,
        "valid_format": True,
        "exists": exists,
        "message": "Ya existe un miembro con esta cédula." if exists else "Cédula disponible.",
    })

@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_dar_salida(request, pk):
    return salida_form(request, pk)
    return salida_form(request, pk)

@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
def nuevo_creyente_dar_salida(request, pk):
    return salida_form(request, pk)


@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def nuevo_creyente_detalle(request, pk):
    """
    Detalle del Nuevo Creyente.
    Usa la plantilla basada en la ficha de miembro, pero enfocada a seguimiento.
    """
    miembro = get_object_or_404(Miembro, pk=pk, nuevo_creyente=True)

    # Parámetro global (lo usa la plantilla de detalle tipo miembro)
    edad_minima = get_edad_minima_miembro_oficial()

    # Expediente (si existe)
    expediente = (
        NuevoCreyenteExpediente.objects
        .filter(miembro=miembro)
        .select_related("responsable")
        .first()
    )
    modulo_nuevo_creyente_activo = _modulo_nuevo_creyente_activo()
    estado_exp = (getattr(expediente, "estado", "") or "").strip().lower()
    expediente_abierto = bool(expediente and estado_exp == "abierto")

    expediente_abierto = bool(expediente and getattr(expediente, "estado", None) == "abierto")

    context = {
        "miembro": miembro,
        "expediente": expediente,
        "expediente_abierto": expediente_abierto,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
         "modulo_nuevo_creyente_activo": modulo_nuevo_creyente_activo,
             "expediente": expediente,
        "expediente_abierto": expediente_abierto,
    }

    # ✅ Ajusta este nombre si tu plantilla se llama distinto
    return render(request, "miembros_app/nuevo_creyente_detalle.html", context)



@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def miembro_inactivo_detalle(request, pk):
    """
    Pantalla única para ver el resumen de salida de un registro inactivo.
    Funciona tanto para Miembro como para Nuevo Creyente (porque ambos viven en Miembro).
    """
    miembro = get_object_or_404(Miembro, pk=pk)

    # Si por alguna razón el registro está activo, lo mandamos a su detalle real
    if miembro.activo:
        if getattr(miembro, "nuevo_creyente", False):
            return redirect("miembros_app:nuevo_creyente_detalle", pk=miembro.pk)
        return redirect("miembros_app:detalle", pk=miembro.pk)

    hoy = timezone.localdate()

    dias_desde_salida = None
    if miembro.fecha_salida:
        try:
            dias_desde_salida = (hoy - miembro.fecha_salida).days
        except Exception:
            dias_desde_salida = None

    context = {
        "miembro": miembro,
        "hoy": hoy,
        "dias_desde_salida": dias_desde_salida,
        "es_nuevo_creyente": bool(getattr(miembro, "nuevo_creyente", False)),
    }
    return render(request, "miembros_app/inactivo_detalle.html", context)



@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
@require_http_methods(["GET", "POST"])
def reincorporar_miembro(request, pk):
    miembro = get_object_or_404(Miembro, pk=pk)

    # Solo aplica para inactivos
    if miembro.activo:
        messages.info(request, "Este registro ya está activo.")
        if getattr(miembro, "nuevo_creyente", False):
            return redirect("miembros_app:nuevo_creyente_detalle", pk=miembro.pk)
        return redirect("miembros_app:detalle", pk=miembro.pk)

    # =========================
    # SNAPSHOT (ANTES de tocar)
    # =========================
    estado_antes = getattr(miembro, "estado_miembro", "") or ""
    etapa_antes = getattr(miembro, "etapa_actual", "") or ""
    razon_txt = "—"
    if getattr(miembro, "razon_salida", None):
        # nombre es el __str__ también, pero así queda claro
        razon_txt = getattr(miembro.razon_salida, "nombre", None) or str(miembro.razon_salida)

    es_nuevo_creyente = bool(getattr(miembro, "nuevo_creyente", False))

    # Para UI: si la razón permite carta, lo mostramos como “requiere”
    requiere_carta = bool(
        miembro.razon_salida and getattr(miembro.razon_salida, "permite_carta", False)
    )

    if request.method == "POST":
        form = MiembroReingresoForm(request.POST)  # ✅ sin instance

        if form.is_valid():
            # ✅ Reactivar
            miembro.activo = True
            miembro.fecha_reingreso = timezone.localdate()

            # ✅ Limpiar salida
            miembro.razon_salida = None
            miembro.fecha_salida = None
            miembro.comentario_salida = ""

            # ✅ Tu lógica automática (la que ya tienes)
            if es_nuevo_creyente:
                miembro.nuevo_creyente = True
                miembro.estado_miembro = "observacion"
                miembro.origen_reingreso = "descarriado"
                miembro.estado_pastoral_reingreso = "reconciliado"
            else:
                miembro.nuevo_creyente = False
                miembro.estado_miembro = "observacion"

                if estado_antes == "descarriado":
                    miembro.origen_reingreso = "descarriado"
                    miembro.estado_pastoral_reingreso = "reconciliado"
                elif estado_antes == "trasladado":
                    miembro.origen_reingreso = "traslado"
                    miembro.estado_pastoral_reingreso = "integrado"
                else:
                    miembro.estado_pastoral_reingreso = (
                        miembro.estado_pastoral_reingreso or "observacion"
                    )

            miembro.save()

            # ✅ Bitácora
            miembro.log_event(
                tipo="reingreso",
                titulo="Reincorporación registrada",
                detalle=f"Razón de salida anterior: {razon_txt}",
                user=request.user,
                estado_from=estado_antes,
                estado_to=getattr(miembro, "estado_miembro", "") or "",
                etapa_from=etapa_antes,
                etapa_to=getattr(miembro, "etapa_actual", "") or "",
            )

            if miembro.nuevo_creyente:
                messages.success(request, "Reingreso registrado: Nuevo Creyente (seguimiento).")
                return redirect("miembros_app:nuevo_creyente_detalle", pk=miembro.pk)

            messages.success(request, "Reincorporación registrada: Miembro en observación.")
            return redirect("miembros_app:detalle", pk=miembro.pk)

        messages.error(request, "Hay errores en el formulario.")
    else:
        form = MiembroReingresoForm()  # ✅ sin instance

    return render(
        request,
        "miembros_app/reincorporacion_form.html",
        {
            "miembro": miembro,
            "form": form,
            "requiere_carta": requiere_carta,
            "es_nuevo_creyente": es_nuevo_creyente,
            "estado_anterior": estado_antes,  # para que la plantilla explique lo que hará
        }
    )


@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
@require_http_methods(["GET", "POST"])
def salida_form(request, pk):

    miembro = get_object_or_404(Miembro, pk=pk)

    from nuevo_creyente_app.models import NuevoCreyenteExpediente

    # Bloqueo: no permitir salida si tiene expediente abierto
    tiene_expediente_abierto = NuevoCreyenteExpediente.objects.filter(
        miembro=miembro,
        estado__iexact="abierto"
    ).exists()

    if tiene_expediente_abierto:
        messages.error(
            request,
            "No se puede dar salida a este nuevo creyente porque "
            "tiene un expediente de seguimiento ABIERTO. "
            "Cierra primero el expediente."
        )
        return redirect("miembros_app:nuevo_creyente_detalle", pk=miembro.pk)

    estado_antes = getattr(miembro, "estado_miembro", "") or ""
    etapa_antes = getattr(miembro, "etapa_actual", "") or ""


    # Permisos (igual que tu vista actual)
    if not (request.user.is_superuser or request.user.has_perm("miembros_app.change_miembro")):
        return HttpResponseForbidden("No tienes permisos para dar salida a miembros.")

    # Si ya está inactivo, no repetir
    if not miembro.activo:
        messages.info(request, "Este registro ya está inactivo. No se puede registrar una salida de nuevo.")
        return redirect("miembros_app:detalle", pk=miembro.pk)

    # ✅ Regla: si es nuevo creyente y tiene expediente ABIERTO, no permitir salida
    es_nuevo_creyente = bool(getattr(miembro, "nuevo_creyente", False))
    if es_nuevo_creyente:
        expediente_abierto = NuevoCreyenteExpediente.objects.filter(
            miembro=miembro
        ).exclude(estado="cerrado").first()
        if expediente_abierto:
            messages.error(
                request,
                "Este nuevo creyente está en seguimiento. Primero debes cerrar el expediente desde el módulo Nuevo Creyente."
            )
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            return redirect("miembros_app:nuevo_creyente_detalle", pk=miembro.pk)

    if request.method == "POST":
        form = MiembroSalidaForm(request.POST, instance=miembro)
        if form.is_valid():
            miembro_editado = form.save(commit=False)

            # Marcar inactivo
            miembro_editado.activo = False

            # Fecha por defecto
            if not miembro_editado.fecha_salida:
                miembro_editado.fecha_salida = timezone.localdate()

            # ✅ Estado al dar salida (según razón configurada)
            if miembro_editado.razon_salida and miembro_editado.razon_salida.estado_resultante:
                miembro_editado.estado_miembro = miembro_editado.razon_salida.estado_resultante


            miembro_editado.save()
            

            miembro.log_event(
                tipo="salida",
                titulo="Salida registrada",
                detalle=f"Razón: {miembro_editado.razon_salida or '—'}",
                user=request.user,
                estado_from=estado_antes,
                estado_to=getattr(miembro_editado, "estado_miembro", "") or "",
                etapa_from=etapa_antes,
                etapa_to=getattr(miembro_editado, "etapa_actual", "") or "",
)

            messages.success(request, "Salida registrada correctamente. El registro ha quedado inactivo.")
            return redirect("miembros_app:inactivo_detalle", pk=miembro_editado.pk)

        messages.error(request, "Hay errores en el formulario. Revisa los campos marcados.")
    else:
        form = MiembroSalidaForm(instance=miembro)

    return render(
        request,
        "miembros_app/salida_form.html",
        {
            "miembro": miembro,
            "form": form,
            "es_nuevo_creyente": es_nuevo_creyente,
        },
    )

# Agregar esta vista a views.py de miembros_app

@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_bitacora_add(request, pk):
    """
    Agregar una entrada manual a la bitácora del miembro.
    """
    from .models import MiembroBitacora
    
    miembro = get_object_or_404(Miembro, pk=pk)
    
    tipo = request.POST.get("tipo", "sistema")
    texto = request.POST.get("texto", "").strip()
    
    if texto:
        # Determinar título según tipo
        if tipo == "nota":
            titulo = "Nota registrada"
        else:
            titulo = "Mensaje"
        
        MiembroBitacora.objects.create(
            miembro=miembro,
            tipo=tipo,
            titulo=titulo,
            detalle=texto,
            creado_por=request.user,
        )
        
        messages.success(request, "Entrada agregada a la bitácora.")
    else:
        messages.error(request, "El texto no puede estar vacío.")
    
    return redirect("miembros_app:detalle", pk=miembro.pk)


# ============================================
# Agregar esta URL en urls.py de miembros_app:
# ============================================
# path('miembros/<int:pk>/bitacora/add/', views.miembro_bitacora_add, name='bitacora_add'),

@login_required
@require_POST
@permission_required("miembros_app.change_miembro", raise_exception=True)
def miembro_bitacora_add(request, pk):
    """
    Agregar una entrada manual a la bitácora del miembro.
    """
    from .models import MiembroBitacora
    
    miembro = get_object_or_404(Miembro, pk=pk)
    
    tipo = request.POST.get("tipo", "sistema")
    texto = request.POST.get("texto", "").strip()
    
    if texto:
        # Determinar título según tipo
        if tipo == "nota":
            titulo = "Nota registrada"
        else:
            titulo = "Mensaje"
        
        MiembroBitacora.objects.create(
            miembro=miembro,
            tipo=tipo,
            titulo=titulo,
            detalle=texto,
            creado_por=request.user,
        )
        
        messages.success(request, "Entrada agregada a la bitácora.")
    else:
        messages.error(request, "El texto no puede estar vacío.")
    
    return redirect("miembros_app:detalle", pk=miembro.pk)


# ============================================
# Agregar esta URL en urls.py de miembros_app:
# ============================================
# path('miembros/<int:pk>/bitacora/add/', views.miembro_bitacora_add, name='bitacora_add'),

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def validar_telefono(request):

    telefono = request.GET.get("telefono", "")
    pk = request.GET.get("pk", "")  # para edición

    # Normalizar: solo dígitos, quitar el 1 inicial si tiene 11
    telefono_norm = re.sub(r"\D+", "", telefono)
    if len(telefono_norm) == 11 and telefono_norm.startswith("1"):
        telefono_norm = telefono_norm[1:]
    telefono_norm = telefono_norm[:10]

    if not telefono_norm:
        return JsonResponse({"existe": False})

    qs = Miembro.objects.filter(telefono_norm=telefono_norm)

    if pk:
        qs = qs.exclude(pk=pk)

    return JsonResponse({"existe": qs.exists()})



from .models import Miembro, ZonaGeo


@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def mapa_miembros(request):
    """
    Pantalla del mapa (Leaflet). Los puntos se cargan por AJAX desde la API.
    """
    return render(request, "miembros_app/mapa_miembros.html")


@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def api_mapa_miembros(request):
    """
    Devuelve puntos por zona (sector/ciudad/provincia) con conteo,
    usando ZonaGeo como cache de coordenadas.
    """
    # Filtros
    activo = request.GET.get("activo", "")  # "1" / "0" / ""
    tipo = request.GET.get("tipo", "")      # "miembro" / "nuevo" / ""
    estado = request.GET.get("estado", "")  # activo/pasivo/...

    qs = Miembro.objects.all()

    if activo == "1":
        qs = qs.filter(activo=True)
    elif activo == "0":
        qs = qs.filter(activo=False)

    if tipo == "miembro":
        qs = qs.filter(nuevo_creyente=False)
    elif tipo == "nuevo":
        qs = qs.filter(nuevo_creyente=True)

    if estado:
        qs = qs.filter(estado_miembro=estado)

    # Agrupar por zona
    agrupado = (
        qs.values("sector", "ciudad", "provincia")
          .annotate(total=Count("id"))
          .order_by("-total")
    )

    puntos = []
    faltantes = []

    for row in agrupado:
        sector = (row.get("sector") or "").strip()
        ciudad = (row.get("ciudad") or "").strip()
        provincia = (row.get("provincia") or "").strip()
        total = row["total"]

        zona = ZonaGeo.objects.filter(
            sector=sector,
            ciudad=ciudad,
            provincia=provincia,
        ).first()

        if zona and zona.lat is not None and zona.lng is not None:
            puntos.append({
                "sector": sector,
                "ciudad": ciudad,
                "provincia": provincia,
                "total": total,
                "lat": zona.lat,
                "lng": zona.lng,
            })
        else:
            # Para que sepas qué zonas no tienen coordenadas todavía
            faltantes.append({
                "sector": sector,
                "ciudad": ciudad,
                "provincia": provincia,
                "total": total,
            })

    return JsonResponse({
        "puntos": puntos,
        "faltantes": faltantes[:50],  # limitamos para no explotar la respuesta
        "faltantes_total": len(faltantes),
    })



@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
@require_POST
def padre_espiritual_add_simple(request, miembro_id):
    miembro = get_object_or_404(Miembro, pk=miembro_id)

    padre_id = request.POST.get("padre_espiritual_id")
    if not padre_id:
        messages.error(request, "Debes seleccionar un padre espiritual.")
        return redirect(
            request.POST.get(
                "return_to",
                reverse("miembros_app:nuevo_creyente_detalle", kwargs={"pk": miembro.pk})
            )
        )

    padre = get_object_or_404(Miembro, pk=int(padre_id))

    if padre == miembro:
        messages.error(request, "Un miembro no puede ser su propio padre espiritual.")
        return redirect(
            request.POST.get(
                "return_to",
                reverse("miembros_app:nuevo_creyente_detalle", kwargs={"pk": miembro.pk})
            )
        )

    miembro.padres_espirituales.add(padre)

    messages.success(
        request,
        f"Padre espiritual asignado: {padre.nombres} {padre.apellidos}."
    )

    return redirect(
        request.POST.get(
            "return_to",
            reverse("miembros_app:nuevo_creyente_detalle", kwargs={"pk": miembro.pk})
        )
    )

@login_required
@permission_required("miembros_app.change_miembro", raise_exception=True)
@require_POST
def padre_espiritual_remove_simple(request, miembro_id, padre_id):
    miembro = get_object_or_404(Miembro, pk=miembro_id)
    padre = get_object_or_404(Miembro, pk=padre_id)

    miembro.padres_espirituales.remove(padre)

    messages.success(
        request,
        f"Padre espiritual removido: {padre.nombres} {padre.apellidos}."
    )

    return redirect(
        request.POST.get(
            "return_to",
            reverse("miembros_app:nuevo_creyente_detalle", kwargs={"pk": miembro.pk})
        )
    )

"""
═══════════════════════════════════════════════════════════════════════════════
CÓDIGO PARA LA VISTA - Organizar relaciones familiares por categoría
═══════════════════════════════════════════════════════════════════════════════

Agregar este código en la vista de detalle del miembro (miembro_detalle_view).

Reemplaza la lógica actual que pasa 'relaciones_familia' y 'relaciones_inferidas'
al contexto por esta nueva lógica que organiza todo en 4 categorías.
═══════════════════════════════════════════════════════════════════════════════
"""

def obtener_relaciones_organizadas(miembro):
    """
    Obtiene todas las relaciones del miembro organizadas en 4 categorías:
    - familia_nuclear: cónyuge e hijos
    - familia_origen: padres y hermanos
    - familia_extendida: abuelos, nietos, tíos, sobrinos, primos, bisabuelos, bisnietos
    - familia_politica: suegros, cuñados, yernos/nueras, consuegros
    
    Retorna un dict con las 4 listas.
    """
    from django.db.models import Q
    from miembros_app.models import MiembroRelacion
    
    mi_id = miembro.id
    
    # Obtener TODAS las relaciones donde aparece este miembro
    relaciones_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro_id=mi_id) | Q(familiar_id=mi_id))
        .select_related("miembro", "familiar")
    )
    
    # Normalizar relaciones
    todas_relaciones = []
    
    for rel in relaciones_qs:
        if rel.miembro_id == mi_id:
            otro = rel.familiar
            tipo = rel.tipo_relacion
        else:
            otro = rel.miembro
            tipo = MiembroRelacion.inverse_tipo(rel.tipo_relacion, rel.miembro.genero)
        
        todas_relaciones.append({
            "otro": otro,
            "tipo": tipo,
            "tipo_label": MiembroRelacion.label_por_genero(tipo, otro.genero),
            "vive_junto": rel.vive_junto,
            "es_responsable": rel.es_responsable,
            "es_inferida": getattr(rel, 'es_inferida', False),  # Compatible si no existe el campo
            "notas": rel.notas,
        })
    
    # Clasificar en categorías
    TIPOS_NUCLEAR = {"conyuge", "hijo"}
    TIPOS_ORIGEN = {"padre", "madre", "hermano"}
    TIPOS_EXTENDIDA = {"abuelo", "nieto", "tio", "sobrino", "primo", "bisabuelo", "bisnieto"}
    TIPOS_POLITICA = {"suegro", "cunado", "yerno", "consuegro"}
    
    familia_nuclear = []
    familia_origen = []
    familia_extendida = []
    familia_politica = []
    
    # Set para evitar duplicados
    ids_agregados = set()
    
    for rel in todas_relaciones:
        otro_id = rel["otro"].id
        
        # Evitar duplicados
        if otro_id in ids_agregados:
            continue
        ids_agregados.add(otro_id)
        
        tipo = rel["tipo"]
        
        if tipo in TIPOS_NUCLEAR:
            familia_nuclear.append(rel)
        elif tipo in TIPOS_ORIGEN:
            familia_origen.append(rel)
        elif tipo in TIPOS_EXTENDIDA:
            familia_extendida.append(rel)
        elif tipo in TIPOS_POLITICA:
            familia_politica.append(rel)
    
    # Ordenar cada categoría
    def orden_tipo(rel):
        """Ordena por tipo de relación para que se vean agrupados."""
        orden = {
            # Nuclear
            "conyuge": 0,
            "hijo": 1,
            # Origen
            "padre": 0,
            "madre": 1,
            "hermano": 2,
            # Extendida
            "abuelo": 0,
            "bisabuelo": 1,
            "tio": 2,
            "primo": 3,
            "sobrino": 4,
            "nieto": 5,
            "bisnieto": 6,
            # Política
            "suegro": 0,
            "cunado": 1,
            "yerno": 2,
            "consuegro": 3,
        }
        return orden.get(rel["tipo"], 99)
    
    familia_nuclear.sort(key=orden_tipo)
    familia_origen.sort(key=orden_tipo)
    familia_extendida.sort(key=orden_tipo)
    familia_politica.sort(key=orden_tipo)
    
    return {
        "familia_nuclear": familia_nuclear,
        "familia_origen": familia_origen,
        "familia_extendida": familia_extendida,
        "familia_politica": familia_politica,
    }




# ═══════════════════════════════════════════════════════════════════════════════
# VERSIÓN SIMPLIFICADA (si prefieres menos cambios)
# ═══════════════════════════════════════════════════════════════════════════════

def obtener_relaciones_organizadas_simple(miembro):
    """
    Versión que retorna las mismas variables pero desempaquetadas,
    listas para agregar directamente al contexto con **.
    """
    from django.db.models import Q
    from miembros_app.models import MiembroRelacion
    
    mi_id = miembro.id
    
    relaciones_qs = (
        MiembroRelacion.objects
        .filter(Q(miembro_id=mi_id) | Q(familiar_id=mi_id))
        .select_related("miembro", "familiar")
    )
    
    TIPOS_NUCLEAR = {"conyuge", "hijo"}
    TIPOS_ORIGEN = {"padre", "madre", "hermano"}
    TIPOS_EXTENDIDA = {"abuelo", "nieto", "tio", "sobrino", "primo", "bisabuelo", "bisnieto"}
    TIPOS_POLITICA = {"suegro", "cunado", "yerno", "consuegro"}
    
    familia_nuclear = []
    familia_origen = []
    familia_extendida = []
    familia_politica = []
    
    ids_agregados = set()
    
    for rel in relaciones_qs:
        if rel.miembro_id == mi_id:
            otro = rel.familiar
            tipo = rel.tipo_relacion
        else:
            otro = rel.miembro
            tipo = MiembroRelacion.inverse_tipo(rel.tipo_relacion, rel.miembro.genero)
        
        if otro.id in ids_agregados:
            continue
        ids_agregados.add(otro.id)
        
        dato = {
            "otro": otro,
            "tipo": tipo,
            "tipo_label": MiembroRelacion.label_por_genero(tipo, otro.genero),
            "vive_junto": rel.vive_junto,
            "es_responsable": rel.es_responsable,
            "es_inferida": rel.es_inferida,
            "notas": rel.notas,
        }
        
        if tipo in TIPOS_NUCLEAR:
            familia_nuclear.append(dato)
        elif tipo in TIPOS_ORIGEN:
            familia_origen.append(dato)
        elif tipo in TIPOS_EXTENDIDA:
            familia_extendida.append(dato)
        elif tipo in TIPOS_POLITICA:
            familia_politica.append(dato)
    
    # Ordenar
    orden = {
        "conyuge": 0, "hijo": 1,
        "padre": 0, "madre": 1, "hermano": 2,
        "abuelo": 0, "bisabuelo": 1, "tio": 2, "primo": 3, "sobrino": 4, "nieto": 5, "bisnieto": 6,
        "suegro": 0, "cunado": 1, "yerno": 2, "consuegro": 3,
    }
    
    for lista in (familia_nuclear, familia_origen, familia_extendida, familia_politica):
        lista.sort(key=lambda r: orden.get(r["tipo"], 99))
    
    return familia_nuclear, familia_origen, familia_extendida, familia_politica