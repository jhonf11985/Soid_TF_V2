# -*- coding: utf-8 -*-
"""
miembros_app/views/reportes.py
Vistas de reportes: listados, cumpleaños, nuevos del mes, salidas, Excel, etc.
"""

from datetime import date, timedelta
from io import BytesIO

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_GET, require_POST
from django.urls import reverse
from django.utils import timezone
from django.db.models import Q
from django.db.models.functions import ExtractDay
from django.template.loader import render_to_string
from django.conf import settings

from openpyxl import Workbook, load_workbook

from miembros_app.models import Miembro, MiembroRelacion, RazonSalidaMiembro
from miembros_app.forms import EnviarFichaMiembroEmailForm
from core.models import ConfiguracionSistema
from core.utils_config import get_edad_minima_miembro_oficial, get_config
from core.utils_email import enviar_correo_sistema

from .utils import MESES_ES, generar_pdf_desde_html
from .miembros import filtrar_miembros


# ═══════════════════════════════════════════════════════════════════════════════
# PANTALLA PRINCIPAL DE REPORTES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reportes_miembros(request):
    """Pantalla principal de reportes del módulo de miembros."""
    return render(request, "miembros_app/reportes/reportes_home.html", {})


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: LISTADO GENERAL IMPRIMIBLE
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_listado_miembros(request):
    """
    Vista de reporte imprimible profesional con header institucional.
    Aplica los mismos filtros que miembro_lista usando filtrar_miembros().
    """
    CFG = ConfiguracionSistema.load()
    
    miembros_base = Miembro.objects.filter(nuevo_creyente=False)
    miembros, filtros_context = filtrar_miembros(request, miembros_base)
    
    # Contar activos e inactivos
    if isinstance(miembros, list):
        activos_count = sum(1 for m in miembros if m.activo)
        inactivos_count = sum(1 for m in miembros if not m.activo)
    else:
        activos_count = miembros.filter(activo=True).count()
        inactivos_count = miembros.filter(activo=False).count()
    
    # Obtener labels para mostrar en metadatos del reporte
    estado = request.GET.get("estado", "").strip()
    genero_filtro = request.GET.get("genero", "").strip()
    categoria_edad_filtro = request.GET.get("categoria_edad", "").strip()
    bautizado = request.GET.get("bautizado", "").strip()
    
    estado_label = dict(Miembro.ESTADOS_MIEMBRO).get(estado, estado) if estado else ""
    genero_label = dict(Miembro.GENERO_CHOICES).get(genero_filtro, genero_filtro) if genero_filtro else ""
    categoria_label = dict(Miembro.CATEGORIA_EDAD_CHOICES).get(categoria_edad_filtro, categoria_edad_filtro) if categoria_edad_filtro else ""
    
    context = {
        "miembros": miembros,
        "query": filtros_context.get("query", ""),
        "estado": estado_label,
        "genero_filtro": genero_label,
        "categoria_edad_filtro": categoria_label,
        "bautizado": bautizado,
        "activos_count": activos_count,
        "inactivos_count": inactivos_count,
        "CFG": CFG,
    }
    
    return render(request, "miembros_app/reportes/reporte_listado_miembros.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: CUMPLEAÑOS DEL MES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_cumple_mes(request):
    """Reporte imprimible de los cumpleaños de un mes."""
    hoy = timezone.localdate()
    edad_minima = get_edad_minima_miembro_oficial()

    # Mes seleccionado
    mes_str = request.GET.get("mes", "").strip()
    mes = int(mes_str) if mes_str.isdigit() and 1 <= int(mes_str) <= 12 else hoy.month
    anio = hoy.year

    # Flags de filtros
    solo_activos = request.GET.get("solo_activos", "1") == "1"
    solo_oficiales = request.GET.get("solo_oficiales", "0") == "1"

    nombre_mes = MESES_ES.get(mes, "")

    # Base: miembros con fecha de nacimiento en ese mes
    miembros = Miembro.objects.filter(
        fecha_nacimiento__isnull=False,
        fecha_nacimiento__month=mes,
    )

    if solo_activos:
        miembros = miembros.filter(activo=True)

    if solo_oficiales:
        cutoff = hoy - timedelta(days=edad_minima * 365)
        miembros = miembros.filter(fecha_nacimiento__lte=cutoff)

    miembros = (
        miembros
        .annotate(dia=ExtractDay("fecha_nacimiento"))
        .order_by("dia", "apellidos", "nombres")
    )

    # Calcular edad que cumplen
    for m in miembros:
        if m.fecha_nacimiento:
            edad_actual = m.calcular_edad()
            m.edad_que_cumple = edad_actual + 1 if edad_actual is not None else None
        else:
            m.edad_que_cumple = None

    context = {
        "miembros": miembros,
        "mes": mes,
        "anio": anio,
        "nombre_mes": nombre_mes,
        "solo_activos": solo_activos,
        "solo_oficiales": solo_oficiales,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/reportes/cumple_mes.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: MIEMBROS NUEVOS DEL MES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_miembros_nuevos_mes(request):
    """Reporte de miembros que ingresaron a la iglesia en un mes concreto."""
    hoy = timezone.localdate()
    edad_minima = get_edad_minima_miembro_oficial()

    # Mes seleccionado (input type="month" -> YYYY-MM)
    mes_str = request.GET.get("mes", "").strip()

    if mes_str:
        try:
            partes = mes_str.split("-")
            anio = int(partes[0])
            mes = int(partes[1])
            if mes < 1 or mes > 12:
                raise ValueError
        except Exception:
            anio, mes = hoy.year, hoy.month
            mes_str = f"{anio:04d}-{mes:02d}"
    else:
        anio, mes = hoy.year, hoy.month
        mes_str = f"{anio:04d}-{mes:02d}"

    solo_activos = request.GET.get("solo_activos", "1") == "1"
    solo_oficiales = request.GET.get("solo_oficiales", "0") == "1"
    query = request.GET.get("q", "").strip()

    nombre_mes = MESES_ES.get(mes, "")

    # Rango de fechas
    fecha_inicio = date(anio, mes, 1)
    if mes == 12:
        fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)

    miembros = Miembro.objects.filter(
        fecha_ingreso_iglesia__isnull=False,
        fecha_ingreso_iglesia__gte=fecha_inicio,
        fecha_ingreso_iglesia__lte=fecha_fin,
    )

    if solo_activos:
        miembros = miembros.filter(activo=True)

    if solo_oficiales:
        cutoff = hoy - timedelta(days=edad_minima * 365)
        miembros = miembros.filter(
            Q(fecha_nacimiento__isnull=False) & Q(fecha_nacimiento__lte=cutoff)
        )

    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(email__icontains=query) |
            Q(telefono__icontains=query) |
            Q(telefono_secundario__icontains=query)
        )

    miembros = miembros.order_by("fecha_ingreso_iglesia", "apellidos", "nombres")

    for m in miembros:
        m.edad_actual = m.calcular_edad() if hasattr(m, "calcular_edad") else None

    context = {
        "miembros": miembros,
        "mes_str": mes_str,
        "anio": anio,
        "mes": mes,
        "nombre_mes": nombre_mes,
        "solo_activos": solo_activos,
        "solo_oficiales": solo_oficiales,
        "query": query,
        "EDAD_MINIMA_MIEMBRO_OFICIAL": edad_minima,
    }

    return render(request, "miembros_app/reportes/reporte_miembros_nuevos_mes.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: MIEMBROS QUE SE FUERON / SALIDAS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_miembros_salida(request):
    """Reporte de miembros inactivos con filtros por fecha y razón de salida."""
    query = request.GET.get("q", "").strip()
    fecha_desde_str = request.GET.get("fecha_desde", "").strip()
    fecha_hasta_str = request.GET.get("fecha_hasta", "").strip()
    razon_salida_id_str = request.GET.get("razon_salida", "").strip()

    miembros = Miembro.objects.filter(activo=False)

    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(email__icontains=query) |
            Q(telefono__icontains=query) |
            Q(telefono_secundario__icontains=query)
        )

    if fecha_desde_str:
        try:
            miembros = miembros.filter(fecha_salida__gte=date.fromisoformat(fecha_desde_str))
        except ValueError:
            pass

    if fecha_hasta_str:
        try:
            miembros = miembros.filter(fecha_salida__lte=date.fromisoformat(fecha_hasta_str))
        except ValueError:
            pass

    razon_salida_id = None
    razon_salida_obj = None
    if razon_salida_id_str and razon_salida_id_str.isdigit():
        razon_salida_id = int(razon_salida_id_str)
        miembros = miembros.filter(razon_salida_id=razon_salida_id)
        razon_salida_obj = RazonSalidaMiembro.objects.filter(pk=razon_salida_id).first()

    razones_disponibles = RazonSalidaMiembro.objects.filter(activo=True).order_by("orden", "nombre")
    miembros = miembros.order_by("-fecha_salida", "apellidos", "nombres")

    context = {
        "miembros": miembros,
        "query": query,
        "fecha_desde": fecha_desde_str,
        "fecha_hasta": fecha_hasta_str,
        "razones_disponibles": razones_disponibles,
        "razon_salida_id": razon_salida_id,
        "razon_salida_obj": razon_salida_obj,
    }

    return render(request, "miembros_app/reportes/reporte_miembros_salida.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: NUEVOS CREYENTES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_nuevos_creyentes(request):
    """Reporte imprimible de nuevos creyentes."""
    query = request.GET.get("q", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    solo_contacto = request.GET.get("solo_contacto", "") == "1"

    miembros = Miembro.objects.filter(nuevo_creyente=True)

    if query:
        miembros = miembros.filter(
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(telefono__icontains=query) |
            Q(telefono_secundario__icontains=query) |
            Q(email__icontains=query)
        )

    if fecha_desde:
        try:
            miembros = miembros.filter(fecha_conversion__gte=fecha_desde)
        except:
            pass

    if fecha_hasta:
        try:
            miembros = miembros.filter(fecha_conversion__lte=fecha_hasta)
        except:
            pass

    if solo_contacto:
        miembros = miembros.filter(
            Q(telefono__isnull=False, telefono__gt="") |
            Q(telefono_secundario__isnull=False, telefono_secundario__gt="") |
            Q(email__isnull=False, email__gt="")
        )

    miembros = miembros.order_by("-fecha_conversion", "-fecha_creacion", "apellidos", "nombres")

    context = {
        "miembros": miembros,
        "query": query,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "solo_contacto": solo_contacto,
        "hoy": timezone.localdate(),
    }

    return render(request, "miembros_app/reportes/reporte_nuevos_creyentes.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE: RELACIONES FAMILIARES
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def reporte_relaciones_familiares(request):
    """
    Reporte: Familias de la Iglesia
    - Familias Extendidas
    - Familias Nucleares
    - Parejas
    """
    from collections import defaultdict
    
    hoy = timezone.localdate()
    CFG = get_config()
    
    query = request.GET.get("q", "").strip()
    tipo_filtro = request.GET.get("tipo", "").strip()
    
    relaciones_qs = (
        MiembroRelacion.objects
        .select_related("miembro", "familiar")
        .filter(miembro__activo=True, familiar__activo=True)
    )
    
    if query:
        relaciones_qs = relaciones_qs.filter(
            Q(miembro__nombres__icontains=query) |
            Q(miembro__apellidos__icontains=query) |
            Q(familiar__nombres__icontains=query) |
            Q(familiar__apellidos__icontains=query)
        )
    
    # Construir grafo de conexiones
    conexiones = defaultdict(set)
    relaciones_info = {}
    
    for rel in relaciones_qs:
        mid, fid = rel.miembro_id, rel.familiar_id
        if mid == fid:
            continue
        
        conexiones[mid].add(fid)
        conexiones[fid].add(mid)
        
        relaciones_info[(mid, fid)] = {
            "tipo": rel.tipo_relacion,
            "vive_junto": rel.vive_junto,
            "es_responsable": rel.es_responsable,
        }
        
        tipo_inverso = MiembroRelacion.inverse_tipo(rel.tipo_relacion, rel.miembro.genero)
        relaciones_info[(fid, mid)] = {
            "tipo": tipo_inverso,
            "vive_junto": rel.vive_junto,
            "es_responsable": False,
        }
    
    # Encontrar grupos familiares (componentes conectados)
    visitados = set()
    grupos = []
    
    for persona_id in list(conexiones.keys()):
        if persona_id in visitados:
            continue
        
        grupo = set()
        cola = [persona_id]
        
        while cola:
            actual = cola.pop(0)
            if actual in visitados:
                continue
            visitados.add(actual)
            grupo.add(actual)
            
            for conectado in conexiones[actual]:
                if conectado not in visitados:
                    cola.append(conectado)
        
        if len(grupo) >= 2:
            grupos.append(grupo)
    
    # Cargar datos de miembros
    todos_ids = set()
    for grupo in grupos:
        todos_ids.update(grupo)
    
    miembros_map = {}
    if todos_ids:
        miembros_map = {m.id: m for m in Miembro.objects.filter(id__in=todos_ids)}
    
    # Funciones auxiliares
    def rel_tipo(a, b):
        info = relaciones_info.get((a, b))
        return info["tipo"] if info else None
    
    def son_conyuges(a, b):
        return rel_tipo(a, b) == "conyuge" or rel_tipo(b, a) == "conyuge"
    
    def es_hijo_de(hijo_id, padre_id):
        return rel_tipo(padre_id, hijo_id) == "hijo" or rel_tipo(hijo_id, padre_id) in ("padre", "madre")
    
    def obtener_parejas_en_grupo(grupo_ids):
        parejas = []
        vistos = set()
        for id1 in grupo_ids:
            for id2 in grupo_ids:
                if id1 >= id2:
                    continue
                if son_conyuges(id1, id2) and (id1, id2) not in vistos:
                    parejas.append((id1, id2))
                    vistos.add((id1, id2))
        return parejas
    
    # Clasificar grupos (simplificado)
    familias_extendidas = []
    familias_nucleares = []
    parejas = []
    
    for grupo_ids in grupos:
        miembros_grupo = [miembros_map[mid] for mid in grupo_ids if mid in miembros_map]
        if len(miembros_grupo) < 2:
            continue
        
        parejas_grupo = obtener_parejas_en_grupo(grupo_ids)
        
        # Determinar apellido principal
        apellidos_count = defaultdict(int)
        for m in miembros_grupo:
            apellidos_count[m.apellidos] += 1
        apellido_principal = max(apellidos_count, key=apellidos_count.get)
        
        # Solo pareja (sin hijos)
        if len(miembros_grupo) == 2 and parejas_grupo:
            p1_id, p2_id = parejas_grupo[0]
            p1 = miembros_map.get(p1_id)
            p2 = miembros_map.get(p2_id)
            
            if p1 and p2:
                parejas.append({
                    "persona1": {"id": p1.id, "nombre_completo": f"{p1.nombres} {p1.apellidos}"},
                    "persona2": {"id": p2.id, "nombre_completo": f"{p2.nombres} {p2.apellidos}"},
                })
        elif len(miembros_grupo) > 2:
            # Familia nuclear o extendida
            familias_nucleares.append({
                "apellido": apellido_principal,
                "miembros": [{"nombre_completo": f"{m.nombres} {m.apellidos}"} for m in miembros_grupo],
                "total_miembros": len(miembros_grupo),
            })
    
    # Aplicar filtro por tipo
    if tipo_filtro == "extendida":
        familias_nucleares = []
        parejas = []
    elif tipo_filtro == "nuclear":
        familias_extendidas = []
        parejas = []
    elif tipo_filtro == "pareja":
        familias_extendidas = []
        familias_nucleares = []
    
    context = {
        "familias_extendidas": familias_extendidas,
        "familias_nucleares": familias_nucleares,
        "parejas": parejas,
        "total_familias_extendidas": len(familias_extendidas),
        "total_familias_nucleares": len(familias_nucleares),
        "total_parejas": len(parejas),
        "query": query,
        "tipo_filtro": tipo_filtro,
        "hoy": hoy,
        "CFG": CFG,
    }
    
    return render(request, "miembros_app/reportes/reporte_relaciones_familiares.html", context)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR E IMPORTAR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@require_GET
@permission_required("miembros_app.view_miembro", raise_exception=True)
def exportar_miembros_excel(request):
    """Exporta a Excel los miembros que se están viendo en el listado."""
    miembros_base = Miembro.objects.filter(nuevo_creyente=False)
    miembros, filtros_context = filtrar_miembros(request, miembros_base)

    wb = Workbook()
    ws = wb.active
    ws.title = "Miembros"

    headers = [
        "ID", "Nombres", "Apellidos", "Edad", "Género", "Estado",
        "Categoría edad", "Teléfono", "Email", "Bautizado", "Fecha ingreso", "Activo",
    ]
    ws.append(headers)

    for m in miembros:
        try:
            edad = m.edad
        except Exception:
            edad = None

        genero_display = m.get_genero_display() if hasattr(m, "get_genero_display") else m.genero
        estado_display = m.get_estado_miembro_display() if hasattr(m, "get_estado_miembro_display") else m.estado_miembro
        categoria_display = m.get_categoria_edad_display() if hasattr(m, "get_categoria_edad_display") else m.categoria_edad
        fecha_ingreso_str = m.fecha_ingreso_iglesia.strftime("%d/%m/%Y") if m.fecha_ingreso_iglesia else ""

        row = [
            m.id,
            m.nombres,
            m.apellidos,
            edad if edad is not None else "",
            genero_display or "",
            estado_display or "",
            categoria_display or "",
            m.telefono or "",
            m.email or "",
            "Sí" if m.bautizado_confirmado else "No",
            fecha_ingreso_str,
            "Activo" if m.activo else "Inactivo",
        ]
        ws.append(row)

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="miembros_filtrados.xlsx"'
    return response


@login_required
@require_POST
@permission_required("miembros_app.add_miembro", raise_exception=True)
def importar_miembros_excel(request):
    """Importa miembros desde un archivo de Excel."""
    from datetime import datetime
    
    archivo = request.FILES.get("archivo_excel")

    if not archivo:
        messages.error(request, "No se ha enviado ningún archivo de Excel.")
        return redirect("miembros_app:lista")

    try:
        wb = load_workbook(filename=archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"El archivo no parece ser un Excel válido: {e}")
        return redirect("miembros_app:lista")

    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        messages.error(request, "El archivo está vacío.")
        return redirect("miembros_app:lista")

    header_map = {str(name).strip().lower(): idx for idx, name in enumerate(header_row) if name}

    columnas_obligatorias = ["nombres", "apellidos"]
    faltantes = [col for col in columnas_obligatorias if col not in header_map]

    if faltantes:
        messages.error(request, f"Faltan columnas obligatorias: {', '.join(faltantes)}")
        return redirect("miembros_app:lista")

    mapa_genero = {
        "masculino": "masculino", "m": "masculino", "hombre": "masculino", "male": "masculino",
        "femenino": "femenino", "f": "femenino", "mujer": "femenino", "female": "femenino",
    }

    mapa_estado = {
        "activo": "activo", "pasivo": "pasivo",
        "en observación": "observacion", "en observacion": "observacion", "observacion": "observacion",
        "disciplina": "disciplina", "en disciplina": "disciplina",
        "descarriado": "descarriado", "catecumeno": "catecumeno", "catecúmeno": "catecumeno",
    }

    creados = omitidos = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all((cell is None or str(cell).strip() == "") for cell in row):
            continue

        def get_value(nombre_col):
            idx = header_map.get(nombre_col)
            if idx is None:
                return ""
            value = row[idx]
            return "" if value is None else str(value).strip()

        nombres = get_value("nombres")
        apellidos = get_value("apellidos")

        if not nombres and not apellidos:
            omitidos += 1
            continue

        genero_val = mapa_genero.get(get_value("genero").lower(), "") if "genero" in header_map else ""
        estado_val = mapa_estado.get(get_value("estado").lower(), "activo") if "estado" in header_map else "activo"
        telefono = get_value("telefono")
        email = get_value("email")

        fecha_nacimiento = None
        if "fecha_nacimiento" in header_map:
            valor_fecha = row[header_map["fecha_nacimiento"]]
            if valor_fecha:
                if hasattr(valor_fecha, "year"):
                    try:
                        fecha_nacimiento = valor_fecha.date()
                    except AttributeError:
                        fecha_nacimiento = valor_fecha
                else:
                    texto_fecha = str(valor_fecha).strip()
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                        try:
                            fecha_nacimiento = datetime.strptime(texto_fecha, fmt).date()
                            break
                        except ValueError:
                            continue

        try:
            Miembro.objects.create(
                nombres=nombres,
                apellidos=apellidos,
                genero=genero_val,
                telefono=telefono,
                email=email,
                fecha_nacimiento=fecha_nacimiento,
                estado_miembro=estado_val,
            )
            creados += 1
        except Exception:
            omitidos += 1

    messages.success(request, f"Importación completada. Creados: {creados}. Omitidos: {omitidos}.")
    return redirect("miembros_app:lista")


# ═══════════════════════════════════════════════════════════════════════════════
# ENVIAR LISTADO POR EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@permission_required("miembros_app.view_miembro", raise_exception=True)
def listado_miembros_enviar_email(request):
    """Genera un PDF del listado de miembros y lo envía por correo."""
    if request.method == "POST":
        form = EnviarFichaMiembroEmailForm(request.POST)
        if form.is_valid():
            destinatario = form.cleaned_data["destinatario"]
            asunto = form.cleaned_data.get("asunto") or "Listado de miembros - SOID"
            mensaje = form.cleaned_data.get("mensaje") or ""

            try:
                miembros = Miembro.objects.filter(activo=True)
                
                q = request.GET.get('q', '').strip()
                if q:
                    miembros = miembros.filter(
                        Q(nombres__icontains=q) |
                        Q(apellidos__icontains=q) |
                        Q(cedula__icontains=q)
                    )
                
                html_string = render_to_string(
                    "miembros_app/reportes/reporte_listado_miembros.html",
                    {"miembros": miembros, "modo_pdf": True},
                    request=request
                )
                pdf_bytes = generar_pdf_desde_html(html_string)

                enviar_correo_sistema(
                    asunto=asunto,
                    mensaje=mensaje,
                    destinatarios=[destinatario],
                    adjuntos=[("listado_miembros.pdf", pdf_bytes)],
                )

                messages.success(request, f"Correo enviado a {destinatario}")
                return redirect("miembros_app:reporte_listado_miembros")

            except Exception as e:
                messages.error(request, f"Error al enviar: {e}")
                return redirect("miembros_app:reporte_listado_miembros")
    else:
        form = EnviarFichaMiembroEmailForm()

    return render(request, "miembros_app/email/enviar_listado_form.html", {
        "form": form,
        "titulo": "Enviar listado de miembros por correo",
        "descripcion": "Se generará un PDF del listado actual.",
        "url_cancelar": reverse("miembros_app:reporte_listado_miembros"),
        "adjunto_auto_nombre": "listado_miembros.pdf",
    })